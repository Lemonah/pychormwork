# -*- coding: utf-8 -*-
# @Time : 2021/3/5 20:27
# @Author :PANZER
# @Email : 453989453@a163.com
# @File : jb_call.py
# @Project : 基本呼叫

import yaml
from loguru import logger  # 日志
from concurrent.futures.thread import ThreadPoolExecutor  #  线程池
from tqdm import tqdm
import os, random, openpyxl, jsonpath, time
from openpyxl.styles import Border, Side, colors, Alignment  # 边框居中设置
from pychormwork.MiguTemplate.businessIntegration.COMPATIBILITY_TEST import test_call


class JbCall:

    # debug_case.yaml  jb_testcase.yaml
    work_path = os.path.dirname(os.path.dirname(__file__))
    # logger.info(f'当前工作路径：{work_path}')
    yaml_file = os.path.join(work_path, r'jb_call\jb_testcase.yaml ')
    time.sleep(2)
    save_path = input('输入《基本呼叫》保存路径: ')
    save_file = os.path.join(save_path, '四川省端到端基本呼叫分析优化-{}({}).xlsx')


    # def __init__(self, yaml_file, save_file):
    #     """
    #
    #     :param yaml_file:
    #     :param save_file:
    #     """
    #     self.yaml_file = yaml_file
    #     self.save_file = save_file

    # 定义边框样式
    def my_border(self):
        """ 定义边框样式 """
        border = Border(top=Side(border_style='thin', color=colors.BLACK),
                        bottom=Side(border_style='thin', color=colors.BLACK),
                        left=Side(border_style='thin', color=colors.BLACK),
                        right=Side(border_style='thin', color=colors.BLACK))
        return border

    # 设置边框
    def set_border(self, ws, column):
        """ 设置边框 """
        all_w = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                 'U', 'V', 'W', 'X', 'Y', 'Z']
        re = all_w.index(column)
        for q in all_w[:re + 1]:
            col = ws[q]
            for cell in col:
                cell.border = self.my_border()  # 调用上面边框设置函数

    # 读取yaml文件
    @test_call
    def read_yaml(self):
        """
        读取yaml文件
        :return: --》[{'test_case01':{}},{'test_case02':{}},{'test_case03':{}}]
        """
        try:
            # logger.info('读取yaml文件--》read_yaml')
            with open(self.yaml_file, encoding='utf-8') as f:
                value = yaml.load(f, Loader=yaml.FullLoader)
                # logger.info('yaml文件读取成功，返回数据：{}'.format(value))
                return value
        except Exception as e:
            logger.error('read_yaml异常：{}'.format(e))

    # 设置保存文件
    @test_call
    def set_save_file(self, thread_list):
        """
        读取模板，保存文件
        :param thread_list: 构造的
        :return:
        """
        # r'D:\桌面\testcase\工单\脚本调试\四川省端到端基本呼叫分析优化-{filename}({time})'
        now_path = os.getcwd()
        logger.debug(f'当前工作路径：{now_path}')
        mode_file01 = os.path.join(now_path, r'test_mode/四川省端到端基本呼叫分析优化-基本呼叫场景测试模板.xlsx')
        mode_file02 = os.path.join(now_path, r'test_mode/四川省端到端基本呼叫分析优化-网络-弱网-重载模板.xlsx')
        mode_file03 = os.path.join(now_path, r'test_mode/四川省端到端基本呼叫分析优化-异常操作场景（被叫早拆线）测试模板.xlsx')
        mode_file04 = os.path.join(now_path, r'test_mode/四川省端到端基本呼叫分析优化-异常操作场景（被叫早摘机）测试模板.xlsx')
        mode_file05 = os.path.join(now_path, r'test_mode/四川省端到端基本呼叫分析优化-异常操作场景（呼叫等待）测试模板.xlsx')
        try:
            # logger.info('创建文件对象和构造存储文件')
            filename = jsonpath.jsonpath(thread_list[1], '$..filename')[0]
            test_name = jsonpath.jsonpath(thread_list[1], '$..test_name')[0]
            # logger.info('filename-->:{}//test_name-->:{}'.format(filename, test_name))
            save_filepath = self.save_file.format(filename, thread_list[0])
            # logger.info('save_filepath-->{}'.format(save_filepath))
            if test_name in ['网络切换', '无线弱覆盖', '无线重载']:
                mode_file = mode_file02
            elif test_name == '基本呼叫':
                mode_file = mode_file01
            elif test_name == '被叫早拆线':
                mode_file = mode_file03
            elif test_name == '被叫早摘机':
                mode_file = mode_file04
            else:
                mode_file = mode_file05
            work_book1 = openpyxl.load_workbook(mode_file)
            work_sheet = work_book1['Sheet1']
            # logger.info('测试结果表格创建对象成功')
            return work_book1, work_sheet, save_filepath
        except Exception as e:
            logger.error('{}创建文件对象异常：{}'.format(test_name, e))

        pass

    # 当前测试用例测试日期列表
    def test_date(self, thread_list):
        date = jsonpath.jsonpath(thread_list[1], '$..testdate')[0]  # [{t1:*,t2:*}]
        time_list = []
        for key in date:
            d_value = date.get(key)
            if d_value is None:
                # logger.info('时间为空跳过')
                pass
            else:
                time_list.append(d_value)
        # logger.info('time_list:{}'.format(time_list))
        return time_list  # [time1,time2,...]

    # 所有地址列表
    def test_address(self, thread_list):
        address_dict = jsonpath.jsonpath(thread_list[1], '$..address')[0]
        address_list = []
        for key in address_dict:
            if address_dict.get(key) is None:
                # logger.info('test_address-->为空,跳过此条地址')
                pass
            else:
                address_list.append(address_dict.get(key))
        # logger.info('address_list:{}'.format(address_list))
        return address_list  # [[d1],[d2],[d3]...]

    # 当前测试用例地址列表
    def test_address_list(self, thread_list):
        time = thread_list[0]  # time
        time_list = self.test_date(thread_list)  # [time1,time2,...]  # 当前用例所有时间
        address_list = self.test_address(thread_list)  # [[d1],[d2],[d3]...] 当前用例所有地址
        num = time_list.index(time)  # 时间下标
        address = address_list[num]  # 根据时间下标获取对应的地址列表
        # logger.info('当前测试用例地址列表-->{}'.format(address))
        return address  # [d1.1,d1.2,d1.3,...]

    # 设置卡顿列表;
    def lag_list(self, test_name, range_num):
        # logger.info(f'获取{test_name}的异常数据列表,当前运行次数为{range_num}')
        lag_list = []
        if test_name == '基本呼叫':  # 5
            lag_num = random.randint(3, 7)
            pass
        elif test_name == '无线重载':  # 8
            lag_num = random.randint(6, 10)
            pass
        elif test_name == '网络切换':  # 6
            lag_num = random.randint(4, 8)
            pass
        elif test_name == '无线弱覆盖':  # 10
            lag_num = random.randint(8, 12)
            pass
        elif test_name == '被叫早拆线':  # 6
            lag_num = random.randint(4, 8)
            pass
        elif test_name == '被叫早摘机':  # 5
            lag_num = random.randint(3, 7)
            pass
        elif test_name == '呼叫等待':  # 7
            lag_num = random.randint(5, 9)
        for i in range(lag_num+1):
            lag_list.append(random.randint(1, range_num))
        return lag_list
        pass

    # 获取测试场景名称
    def set_testname_list(self, value):
        """
        获取测试用例名称集合的列表，用于循环启动多线程
        :param value:   YamlHandle().yaml_read(self.yamlpath)获取的测试用例数据
        :return: test_key_list--》测试场景名称列表
        """
        test_key_list = []
        for test_dict in value:
            for test_key in test_dict:
                test_key_list.append(test_key)
        # logger.info('测试用例集为：{}'.format(test_key_list))
        return test_key_list

    # 线程数的参数
    def set_thread_list(self, i, value, test_key_list):
        """
        :param i: 最外层循环次数
        :param value: yaml文件读取的值
        :param test_key_list: test_case01,test_case02,....
        :return:
        """
        # 取测试日期
        data = jsonpath.jsonpath(value[i], '$..testdate')
        test_date_list = []  # 测试日期列表
        for test_date_key in data[0]:
            value01 = data[0].get(test_date_key)
            if value01 is None:
                pass
            else:
                test_date_list.append(value01)
        # logger.info('当前测试用例集时间：{}'.format(test_date_list))
        # 获取testcase值
        case_data = jsonpath.jsonpath(value[i], '$..{}'.format(test_key_list[i]))
        # logger.info('测试集testcase值：{}'.format(case_data))
        # 组合列表
        thread_list = []
        for v in test_date_list:
            list01 = [v, case_data[0]]
            thread_list.append(list01)
        return thread_list  # [[time,{}],[time2,{}]]

    # 基本呼叫主场景测试
    def set_mainScene_jb(self, run_now, sence_num):
        """

        :param run_now:
        :param sence_num:  步进数
        :return:
        """
        try:
            subScene_list = ['被叫已开通视频彩铃', '被叫未开通视频彩铃']
            if run_now-sence_num / 12 <= 0.5:  # 需要步进修改
                subScene = subScene_list[0]
            else:
                subScene = subScene_list[1]
            return subScene
        except Exception as e:
            logger.error(f'set_mainScene_jb异常：{e}')

    # 设置子场景--呼叫等待 没有子场景
    def set_subScene(self, test_name, run_now, sence_num=0):
        subScene01 = ['主叫在弱覆盖区域', '被叫在弱覆盖区域', '主被叫在弱覆盖区域']
        subScene02 = ['主叫在无线重载区域', '被叫在无线重载区域', '主被叫在无线重载区域']
        subScene03 = ['主叫网络切换', '被叫网络切换']
        subScene04 = ['主叫终端支持视频彩铃', '主叫终端不支持视频彩铃']  # 早拆线,早摘机,基本呼叫场景
        try:
            if test_name == '无线弱覆盖':
                if 1<= run_now % 9 <= 3:
                    subScene = subScene01[0]
                elif 3 < run_now % 9 <= 6:
                    subScene = subScene01[1]
                else:
                    subScene = subScene01[2]
                pass
            elif test_name == '无线重载':
                if 1<= run_now % 9 <= 3:
                    subScene = subScene02[0]
                elif 3 < run_now % 9 <= 6:
                    subScene = subScene02[1]
                else:
                    subScene = subScene02[2]
                pass
            elif test_name == '网络切换':
                if 0 < run_now-sence_num / 6 <= 0.5:  # 需要步进 6
                    subScene = subScene03[0]
                else:
                    subScene = subScene02[2]
                pass
            elif test_name == '被叫早拆线' or '被叫早摘机':
                if run_now % 2 == 0:
                    subScene = subScene04[1]
                else:
                    subScene = subScene04[0]
            else:  # 基本呼叫
                if 1 <= run_now % 6 <= 3:  # 需要步进12
                    subScene = subScene04[0]
                else:
                    subScene = subScene04[1]
            return  subScene
        except Exception as e:
            logger.error('set_subScene【{}】获取子场景异常--》{}'.format(test_name, e))

    # 设置测试用例
    def set_testCase(self, test_name, run_now):
        testCase01 = ['1.被叫已开通视频彩铃', '2.被叫已开通音频彩铃', '3.被叫未开通彩铃']  # -->重载，弱网，网络切换
        testCase02 = '主叫侧未开始播放视频彩铃，被叫已拆线挂断。'  # --》早拆线
        testCase03 = '主叫侧未开始播放视频彩铃，被叫已摘机接通。'  # --》早摘机
        testCase04 = ['1.A呼叫B, 振铃阶段C呼叫B, B接通C的电话, A和B呼叫等待是否正常',
                      '2.A呼叫B, 振铃阶段C呼叫B, B呼叫C, B是否能看到C的视频彩铃']  # --》呼叫等待
        testCase05 = ['1.主叫VoLTE用户（驻留LTE）音频呼叫VoLTE用户（驻留LTE）',
                      '2.主叫VoLTE用户（驻留LTE）音频呼叫VoLTE用户（驻留CS)',
                      '3.主叫VoLTE用户（驻留CS）音频呼叫VoLTE用户（驻留LTE)']
        try:
            if test_name in ['无线弱覆盖', '无线重载', '网络切换']:
                if run_now % 3 == 1:
                    testCase = testCase01[0]
                elif run_now % 3 == 2:
                    testCase = testCase01[1]
                else:
                    testCase = testCase01[2]
                pass
            elif test_name == '被叫早拆线':
                testCase = testCase02
            elif test_name == '被叫早摘机':
                testCase = testCase03
            elif test_name == '基本呼叫':
                if run_now % 3 == 1:
                    testCase = testCase05[0]
                elif run_now % 3 == 2:
                    testCase = testCase05[1]
                else:
                    testCase = testCase05[2]
            else:  # 呼叫等待
                if run_now % 2 != 0:
                    testCase = testCase04[0]
                else:
                    testCase = testCase04[1]
                pass
            return testCase
        except Exception as e:
            logger.error('set_testCase【{}】设置测试用例异常--》{}'.format(test_name, e))

    # 设置主被叫号码
    def set_phonenumber(self, test_name,):
        # logger.info('构造{}主被叫号码'.format(test_name))
        phonenumber_list = [15982094112, 18382722728, 13637932712]
        a = int(random.uniform(-1, 3))
        phonenumbera = None
        phonenumberb = None
        phonenumberc = None
        if test_name in ['无线弱覆盖', '无线重载', '网络切换', '被叫早拆线', '被叫早摘机', '基本呼叫']:
            if a == 0:
                phonenumbera = phonenumber_list[a]
                phonenumberb = phonenumber_list[a+2]
            elif a == 2:
                phonenumbera = phonenumber_list[a]
                phonenumberb = phonenumber_list[a - 2]
            else:
                phonenumbera = phonenumber_list[a]
                phonenumberb = phonenumber_list[a - 1]
            pass
        else:
            if a == 0:
                phonenumbera = phonenumber_list[a]
                phonenumberb = phonenumber_list[a + 2]
                phonenumberc = phonenumber_list[a + 1]
            elif a == 2:
                phonenumbera = phonenumber_list[a]
                phonenumberb = phonenumber_list[a - 2]
                phonenumberc = phonenumber_list[a - 1]
            else:
                phonenumbera = phonenumber_list[a]
                phonenumberb = phonenumber_list[a-1]
                phonenumberc = phonenumber_list[a+1]
            pass
        # logger.info('{}主被叫号码为：{}，{}，{}'.format(test_name, phonenumbera, phonenumberb, phonenumberc))
        return phonenumbera, phonenumberb, phonenumberc

    # 主叫终端列表设置 -- 呼叫等待不用终端 获取各种data前先获取 只获取一次
    def set_callingTerminal_list(self, test_name):
        callingTerminal_list01 = ['HUAWEI P20', '三星Galaxy s8', 'vivo iQOO', '小米9', 'oppo FIND X']
        a = int(random.uniform(-1, 5))
        # logger.info('set_callingTerminal_list-->test_name:{}'.format(test_name))
        if test_name in ['无线弱覆盖', '无线重载', '网络切换']:
            if a == 0:
                callingTerminal_list = [callingTerminal_list01[a+2], callingTerminal_list01[a]]
            elif a == 4:
                callingTerminal_list = [callingTerminal_list01[a - 2], callingTerminal_list01[a]]
            else:
                callingTerminal_list = [callingTerminal_list01[a + 1], callingTerminal_list01[a]]
            # logger.info('set_callingTerminal_list当前线程使用终端为--》：{}'.format(callingTerminal_list))
            return callingTerminal_list
        elif test_name in ['被叫早拆线', '被叫早摘机', '基本呼叫']:
            callingTerminal_list = [callingTerminal_list01[a], '苹果7']
            # logger.info('set_callingTerminal_list当前线程使用终端为--》：{}'.format(callingTerminal_list))
            return callingTerminal_list
        else:
            pass
            # logger.info('set_callingTerminal_list呼叫等待不需要主叫终端')

    # 设置当前使用的终端
    def set_callingTerminal(self,  test_name, run_now, usecasenumber, callingTerminal_list=None, subScene=None):
        # 用例编号确定
        # callingTerminal_list 这个列表里面只会又两个终端型号
        # test_name, run_now, callingTerminal_list, subScene
        try:
            # logger.info('设置【{}】当前写入终端：{}'.format(test_name, callingTerminal_list))
            if test_name in ['无线弱覆盖', '无线重载', '网络切换']:  # 9换
                # logger.info('第{}次{}usecasenumber-->{}'.format(run_now, test_name, usecasenumber))
                if usecasenumber % 2 != 0:
                    callingterminal = callingTerminal_list[0]
                else:
                    callingterminal = callingTerminal_list[1]
                return callingterminal
            elif test_name in ['被叫早拆线', '被叫早摘机']:  # 2换
                if subScene == '主叫终端支持视频彩铃':
                    callingterminal = callingTerminal_list[0]
                else:
                    callingterminal = callingTerminal_list[1]
                return callingterminal
            elif test_name == '基本呼叫':
                if 1 <= run_now % 6 <= 3:
                    callingterminal = callingTerminal_list[0]
                else:
                    callingterminal = callingTerminal_list[1]
                return callingterminal
            else:  # 呼叫等待不换
                pass
                # logger.info('呼叫等待不需要更换终端')
        except Exception as e:
            logger.error('设置【{}】当前写入终端异常{}。'.format(test_name, e))

    # 当前测试用例地址列表
    def set_address_list(self, thread_list):
        try:
            time = thread_list[0]  # time
            time_list = self.test_date(thread_list)  # [time1,time2,...]
            address_list = self.test_address(thread_list)  # [[d1],[d2],[d3]...]
            num = time_list.index(time)
            address = address_list[num]
            return address  # [d1.1,d1.2,d1.3,...]
        except Exception as e:
            logger.error(f'set_address_list异常：{e}')

    # 设置当前写入测试用例地址
    def set_address(self, test_name, usecasenumber, address_list, range_num):
        if test_name in ['无线弱覆盖', '无线重载']:  # 9*400
            if range_num/36 < usecasenumber <= range_num/18:
                address = address_list[0]
            elif range_num/18 < usecasenumber <= range_num/12:
                address = address_list[1]
            elif range_num/12 < usecasenumber:
                address = address_list[2]
            else:
                address = address_list[3]
            pass
        elif test_name == '网络切换':  # 6*400
            if range_num/24 < usecasenumber <= range_num/12:
                address = address_list[0]
            elif range_num/12 < usecasenumber <= range_num/8:
                address = address_list[1]
            elif range_num/8 < usecasenumber:
                address = address_list[2]
            else:
                address = address_list[3]
            pass
        elif test_name in ['被叫早拆线', '被叫早摘机', '呼叫等待']:  # 2*400
            if range_num/8 < usecasenumber <= range_num/4:
                address = address_list[0]
            elif range_num/4 < usecasenumber <= range_num/2.6:
                address = address_list[1]
            elif range_num/2.6 < usecasenumber:
                address = address_list[2]
            else:
                address = address_list[3]
            pass
        else:  # 呼叫等待12*400
            if range_num/48 < usecasenumber <= range_num/24:
                address = address_list[0]
            elif range_num/24 < usecasenumber <= range_num/16:
                address = address_list[1]
            elif range_num/16 < usecasenumber:
                address = address_list[2]
            else:
                address = address_list[3]
        return address

    # 设置彩铃是否正常--弱网，重载，切换使用
    def set_colorRing(self, run_now, ):
        if run_now % 3 == 0:
            colorRing = '/'
        else:
            colorRing = '正常'
        return colorRing

    # 设置运行次数
    def set_rangenum(self, test_name):
        if test_name in ['无线弱覆盖', '无线重载']:
            while True:
                range_num = int(random.uniform(3600, 3802))
                if (range_num-1) % 9 == 0:
                    # logger.info('{}的总运行次数为{}'.format(test_name,range_num))
                    break
        elif test_name == '网络切换':
            while True:
                range_num = int(random.uniform(2400, 2602))
                if (range_num - 1) % 6 == 0:
                    # logger.info('{}的总运行次数为{}'.format(test_name, range_num))
                    break
        elif test_name == '基本呼叫':
            while True:
                range_num = int(random.uniform(4800, 5002))
                if (range_num - 1) % 12 == 0:
                    # logger.info('{}的总运行次数为{}'.format(test_name, range_num))
                    break
        else:
            while True:
                range_num = int(random.uniform(800, 850))
                if (range_num - 1) % 12 == 0:
                    # logger.info('{}的总运行次数为{}'.format(test_name, range_num))
                    break
        return range_num

    # 设置RSRP
    def set_rsrp(self, test_name):
        """
        # RSRP 弱覆盖（-115，-105），重载、网络切换（-95，-85）
        根据名字判断 随机值
        :param test_name:
        :return:
        """
        if test_name in ['无线重载', '网络切换']:
            RSRP = int(random.uniform(-95, -85))
        else:
            RSRP = int(random.uniform(-115, -105))
        return RSRP
        pass

    # 设置SINR
    def set_sinr(self, test_name):
        """
        # SINR 弱覆盖（-30，30）/10,重载、网络切换（160，250）/10
        :param test_name:
        :return:
        """
        if test_name in ['无线重载', '网络切换']:
            SINR = float('%.1f' % (random.uniform(160, 250) / 10))
        else:
            SINR = float('%.1f' % (random.uniform(-30, 30) / 10))
        pass

    # 设置基本呼叫的通话质量
    def set_conversationQuality(self, run_now, lag_list):
        if run_now in lag_list:
            conversationQuality = '有杂音'
            problem = '主叫终端有杂音'
        else:
            conversationQuality = '正常'
            problem = None
        return conversationQuality, problem

    # 早拆线通话质量
    def set_conversationQuality02(self, run_now, lag_list):
        if run_now % 2 == 0:
            conversationQuality = '是'
            problem = None
        else:
            if run_now in lag_list:
                conversationQuality = '否'
                problem = '主叫终端未正常释放'
            else:
                conversationQuality = '是'
                problem = None
        return conversationQuality, problem

    # 呼叫等待问题
    def set_conversationQuality03(self, run_now, lag_list):
        if run_now in lag_list:
            conversationQuality = '否'
            problem = 'A和B呼叫等待失败'
        else:
            conversationQuality = '是'
            problem = None
        return conversationQuality, problem

    # 设置重载问题
    def set_conversationQuality04(self, run_now, lag_list):
        if run_now in lag_list:
            conversationQuality = '有杂音'
            problem = '通话有轻微杂音'
        else:
            conversationQuality = '正常'
            problem = None
        return conversationQuality, problem

    # 无线弱覆盖-重载-网络切换-data
    def rzw_data(self, usecasenumber, test_name, testdate, fnum, tnum, callingTerminal_list, address_list,
               range_num, run_now, sence_num, lag_list=None):
        """
        获取无线弱覆盖-重载-网络切换-data
        :param usecasenumber: interesting 用例编号 根据不同的测试场景设置步进增加1
        :param test_name: str 根据yaml文件的test_name取值
        :param testdate: str 根据yaml文件的testdate取值
        :param fnum: int 主叫号码
        :param tnum: interesting 被叫号码
        :param callingTerminal_list: list 主叫终端列表 [p1,p2]
        :param address_list: list 测试地址 [d1.1,d1.2,d1.3,d1.4]
        :param range_num: int 根据不同场景去随机设置总运行次数
        :param run_now: int 当前运行到第几次
        :param sence_num: interesting  步进不同场景不一样 用于判断子场景等 初始值为0
        :return:
        """
        # 总共17列
        try:
            testingScenarios = test_name  # 测试主场景
            subScene = self.set_subScene(test_name, run_now, sence_num)  # 逢3合并 每3次更换一次 需要run_now判断  # 子场景
            testCase = self.set_testCase(test_name, run_now)  # %3 的余数来确定 需要run_now
            # testdate = '2021-2-2'  # 测试用例集里面传进来
            # fnum = ''  # 在save_function循环前确定 传进来
            # tnum = ''  # 同上
            # callingTerminal_list = self.set_callingTerminal_list(test_name)  # 主叫终端 [1,2] 根据用例编号来确定 调用这个方法前获取
            callingTerminal = self.set_callingTerminal(test_name, run_now, usecasenumber=usecasenumber,
                                                       callingTerminal_list=callingTerminal_list, subScene=subScene)
            tester = '孙振/马智'
            # RSRP 弱覆盖（-115，-105），重载、网络切换（-95，-95）
            RSRP = self.set_rsrp(test_name)
            # SINR 弱覆盖（-30，30）/10,重载、网络切换（160，250）/10
            SINR = float('%.1f' % (random.uniform(-30, 30) / 10))

            callwait = float('%.1f' % (random.uniform(45, 80) / 10))
            address = self.set_address(test_name, usecasenumber, address_list, range_num)  # 传进来一个地址列表 根据测试用例编号确定 需要一个usecasenumber
            colorRing= self.set_colorRing(run_now)  # 3的倍数为'/', 如果在判断异常列表随机数里为异常，如果为异常须在第17列加测试问题描述
            conversationQuality, problem = self.set_conversationQuality04(run_now, lag_list)
            droppedCall = '否'
            data = [usecasenumber, testingScenarios, subScene, testCase, testdate, fnum, tnum, callingTerminal,
                    tester, address, RSRP, SINR, callwait, colorRing, conversationQuality, droppedCall, problem]
            return data
        except Exception as e:
            logger.error('调用rzw_data异常--》{}'.format(e))
        pass

    # 基本呼叫-data
    def jb_data(self, usecasenumber, test_name, testdate, fnum, tnum, callingTerminal_list, address_list,
                range_num, run_now, sence_num, lag_list):
        try:
            # logger.info('构造{}data'.format(test_name))
            mainScene = self.set_mainScene_jb(run_now, sence_num)
            subScene = self.set_subScene(test_name, run_now, sence_num)
            testCase = self.set_testCase(test_name, run_now)
            callingTerminal = self.set_callingTerminal(test_name, run_now, usecasenumber, callingTerminal_list, subScene)
            tester = '孙振/马智'
            address = self.set_address(test_name, usecasenumber, address_list, range_num)
            callwait = float('%.1f' % (random.uniform(40, 90) / 10))
            conversationQuality, problem = self.set_conversationQuality(run_now, lag_list)
            droppedCall = '否'
            data = [usecasenumber, mainScene, subScene, testCase, testdate, fnum, tnum, callingTerminal,
                    tester, address, callwait, conversationQuality, droppedCall, problem]
            # logger.info('第{}次{}测试data：{}'.format(run_now, test_name, data))
            return data
        except Exception as e:
            logger.error('调用jb_data异常：{}'.format(e))

    # 早摘机data
    def zzj_data(self, usecasenumber, test_name, testdate, fnum, tnum, callingTerminal_list, address_list, range_num,
                 run_now, sence_num, lag_list=None):
        # usecasenumber, test_name, testdate, fnum, tnum, callingTerminal_list, address_list,
                                    #range_num, run_now, sence_num
        try:
            # logger.info('zzj_data-->callingTerminal_list:{}'.format(callingTerminal_list))
            testingScenarios = test_name
            subScene = self.set_subScene(test_name, run_now, sence_num)
            testCase = self.set_testCase(test_name, run_now)
            callingTerminal = self.set_callingTerminal(test_name, run_now, usecasenumber, callingTerminal_list, subScene)
            tester = '孙振/马智'
            callwait = float('%.1f' % (random.uniform(30, 40) / 10))
            address = self.set_address(test_name, usecasenumber, address_list, range_num)
            conversationQuality, problem = self.set_conversationQuality(run_now, lag_list)
            droppedCall = '否'
            data = [usecasenumber, testingScenarios, subScene, testCase, testdate, fnum, tnum, callingTerminal,
                    tester, address, callwait, conversationQuality, droppedCall, problem]
            # logger.info('zzj_data：【{}次】{}数据获取成功：{}'.format(run_now, test_name, data))
            return data
        except Exception as e:
            logger.error('zzj_data-->【{}】获取data数据异常：{}'.format(test_name, e))

    # 早拆线data
    def zcx_data(self, usecasenumber, test_name, testdate, fnum, tnum, callingTerminal_list, address_list, range_num,
                 run_now, sence_num, lag_list=None):
        try:
            testingScenarios = test_name
            subScene = self.set_subScene(test_name, run_now, sence_num)
            testCase = self.set_testCase(test_name, run_now)
            callingTerminal = self.set_callingTerminal(test_name, run_now, usecasenumber, callingTerminal_list, subScene)
            tester = '孙振/马智'
            address = self.set_address(test_name, usecasenumber, address_list, range_num)
            conversationQuality, problem = self.set_conversationQuality02(run_now, lag_list)
            data = [usecasenumber, testingScenarios, subScene, testCase, testdate, fnum, tnum, callingTerminal,
                    tester, address, conversationQuality, problem]
            # logger.info('zcx_data：【{}次】{}数据获取成功：{}'.format(run_now, test_name, data))
            return data
        except Exception as e:
            logger.error('zcx_data数据获取异常--》{}'.format(e))

    # 呼叫等待
    def callwait_data(self, usecasenumber, test_name, testdate, pnuma, pnumb, pnumc, address_list, range_num, run_now,
                      sence_num, lag_list=None):
        testingScenarios = test_name
        testCase = self.set_testCase(test_name, run_now)
        tester = '孙振/马智'
        address = self.set_address(test_name, usecasenumber, address_list, range_num)
        conversationQuality, problem = self.set_conversationQuality03(run_now, lag_list)
        calling_list = ['是', '/']
        if run_now % 2 == 0:
            calling = calling_list[0]
        else:
            calling = calling_list[1]
        data = [usecasenumber, testingScenarios, testCase, testdate, pnuma, pnumb, pnumc,
                tester, address, conversationQuality, calling, problem]
        # logger.info('callwait_data：【{}次】{}数据获取成功：{}'.format(run_now, test_name, data))
        return data

    # 重载-弱网-保存方式
    @test_call
    def save_function01(self, thread_list, address_list, range_num, lag_list):  # [time,{}]
        """
        重载-弱网-保存方式
        :param thread_list: list 包含两个元素-->[time,{test_case}]
        :param address_list: list 测试用例地址列表--> [d1.1,d1.2,d1.3,d1.4]
        :param range_num:  int 运行总次数
        :return:
        """
        # widgets = ['Progress: save_function01', Percentage(), ' ', Bar(marker=RotatingMarker('>-=')),
        #            ' ', ETA(), ' ', FileTransferSpeed()]
        # pbar = ProgressBar(widgets=widgets, maxval=100000).start()
        try:
            # logger.info('调用save_function01')
            test_name = jsonpath.jsonpath(thread_list[1], '$..test_name')[0]
            testdate = thread_list[0]
            usecasenumber = 1
            callingTerminal_list = self.set_callingTerminal_list(test_name)
            # address_list = self.set_address_list(thread_list)
            # range_num = self.set_rangenum(test_name)
            fnum, tnum, x = self.set_phonenumber(test_name)
            work_book, work_sheet, save_file = self.set_save_file(thread_list)
            sence_num = 0
            row = 1  # 行
            c = 1  # 列
            for run_now in tqdm(range(1, range_num)):
                row += 1
                data = self.rzw_data(usecasenumber, test_name, testdate, fnum, tnum, callingTerminal_list, address_list,
                                     range_num, run_now, sence_num, lag_list=lag_list)
                # logger.info('第{}次data:{}'.format(run_now, data))
                for i in data:  # 16
                    work_sheet.cell(row=row, column=c, value=i)
                    work_sheet.cell(row=row, column=c).alignment = Alignment(horizontal='center', vertical='center')
                    if c == 17:
                        if run_now % 3 == 0:
                            work_sheet.merge_cells(start_row=run_now - 1, start_column=3, end_row=run_now + 1,
                                                   end_column=3)
                        if run_now % 9 == 0:
                            usecasenumber += 1
                            sence_num += 9
                            work_sheet.merge_cells(start_row=run_now - 7, start_column=2, end_row=run_now + 1,
                                                   end_column=2)
                            work_sheet.merge_cells(start_row=run_now - 7, start_column=1, end_row=run_now + 1,
                                                   end_column=1)
                        c = 1
                    else:
                        c += 1
            #     pbar.update(10 * run_now + 1)
            #     # time.sleep(0.0001)
            # pbar.finish()
            self.set_border(work_sheet, 'Q')
            work_book.save(save_file)
            # logger.info('{}保存成功'.format(test_name))
        except Exception as e:
            logger.error('调用save_function01异常：{}'.format(e))
        finally:
            work_book.save(save_file)

    # 基本呼叫 保存方式
    @test_call
    def save_function02(self, thread_list, address_list, range_num, lag_list):
        # widgets = ['Progress: save_function02', Percentage(), ' ', Bar(marker=RotatingMarker('>-=')),
        #            ' ', ETA(), ' ', FileTransferSpeed()]
        # pbar = ProgressBar(widgets=widgets, maxval=100000).start()
        try:
            # logger.info('调用save_function02')
            test_name = jsonpath.jsonpath(thread_list[1], '$..test_name')[0]
            testdate = thread_list[0]
            usecasenumber = 1
            callingTerminal_list = self.set_callingTerminal_list(test_name)  # 终端列表
            # logger.info('callingTerminal_list-->:{}'.format(callingTerminal_list))
            # address_list = self.set_address_list(thread_list)
            # range_num = self.set_rangenum(test_name)
            fnum, tnum, x = self.set_phonenumber(test_name)
            work_book, work_sheet, save_file = self.set_save_file(thread_list)
            sence_num = 0
            row = 1  # 行
            c = 1  # 列
            for run_now in range(1, range_num):
                row += 1
                data = self.jb_data(usecasenumber, test_name, testdate, fnum, tnum, callingTerminal_list, address_list,
                                    range_num, run_now, sence_num, lag_list=lag_list)
                # logger.info('第{}写入data'.format(run_now))
                for i in data:
                    work_sheet.cell(row=row, column=c, value=i)
                    work_sheet.cell(row=row, column=c).alignment = Alignment(horizontal='center', vertical='center')
                    #logger.info('写入{}行{}列数据：{}'.format(row-1, c, i))
                    if c == 14:
                        if run_now % 3 == 0:
                            work_sheet.merge_cells(start_row=run_now - 1, start_column=3, end_row=run_now + 1,
                                                   end_column=3)
                            # logger.info('合并第3列{}行——》{}行'.format(run_now - 1, run_now + 1))
                        if run_now % 6 == 0:
                            work_sheet.merge_cells(start_row=run_now - 4, start_column=2, end_row=run_now + 1,
                                                   end_column=2)
                            # logger.info('合并第2列{}行——》{}行'.format(run_now - 4, run_now + 1))
                        if run_now % 12 == 0:
                            work_sheet.merge_cells(start_row=run_now - 10, start_column=1, end_row=run_now + 1,
                                                   end_column=1)
                            # logger.info('合并第1列{}行——》{}行'.format(run_now - 10, run_now + 1))
                            usecasenumber += 1
                            sence_num += 12
                        c = 1
                    else:
                        c += 1
            #     pbar.update(10 * run_now + 1)
            #     # time.sleep(0.0001)
            # pbar.finish()

        except Exception as e:
            logger.error('调用save_function02异常：{}'.format(e))
        finally:
            self.set_border(work_sheet, 'N')
            work_book.save(save_file)
            # logger.info('{}保存成功'.format(test_name))

    # 网络切换 保存方式
    @test_call
    def save_function03(self, thread_list, address_list, range_num, lag_list):
        # widgets = ['Progress: save_function03', Percentage(), ' ', Bar(marker=RotatingMarker('>-=')),
        #            ' ', ETA(), ' ', FileTransferSpeed()]
        # pbar = ProgressBar(widgets=widgets, maxval=100000).start()
        try:
            # logger.info('判断为-{}，调用save_function03')
            test_name = jsonpath.jsonpath(thread_list[1], '$..test_name')[0]
            testdate = thread_list[0]
            usecasenumber = 1
            callingTerminal_list = self.set_callingTerminal_list(test_name)
            # address_list = self.set_address_list(thread_list)
            # range_num = self.set_rangenum(test_name)
            fnum, tnum, x = self.set_phonenumber(test_name)
            work_book, work_sheet, save_file = self.set_save_file(thread_list)
            sence_num = 0
            row = 1  # 行
            c = 1  # 列
            for run_now in range(1, range_num):
                row += 1
                data = self.rzw_data(usecasenumber, test_name, testdate, fnum, tnum, callingTerminal_list, address_list,
                range_num, run_now, sence_num, lag_list=lag_list)
                # logger.info('第{}次{}data：{}'.format(run_now, test_name, data))
                for i in data:
                    work_sheet.cell(row=row, column=c, value=i)
                    work_sheet.cell(row=row, column=c).alignment = Alignment(horizontal='center', vertical='center')
                    if c == 17:
                        if run_now % 3 == 0:
                            work_sheet.merge_cells(start_row=run_now - 1, start_column=3, end_row=run_now + 1,
                                                   end_column=3)
                        if run_now % 6 == 0:
                            work_sheet.merge_cells(start_row=run_now - 4, start_column=2, end_row=run_now + 1,
                                                   end_column=2)
                            work_sheet.merge_cells(start_row=run_now - 4, start_column=1, end_row=run_now + 1,
                                                   end_column=1)
                            usecasenumber += 1
                            sence_num += 6
                        c = 1
                    else:
                        c += 1
            #     pbar.update(10 * run_now + 1)
            #     # time.sleep(0.0001)
            # pbar.finish()
            self.set_border(work_sheet, 'Q')
            work_book.save(save_file)
            # logger.info('{}保存成功'.format(test_name))
        except Exception as e:
            logger.error('调用save_function03异常：{}'.format(e))
        finally:
            work_book.save(save_file)

    # 早摘机 保存方式
    @test_call
    def save_function04(self, thread_list, address_list, range_num, lag_list):
        # widgets = ['Progress: save_function04', Percentage(), ' ', Bar(marker=RotatingMarker('>-=')),
        #            ' ', ETA(), ' ', FileTransferSpeed()]
        # pbar = ProgressBar(widgets=widgets, maxval=100000).start()
        try:
            # logger.info('判断为【被叫早摘机】调用save_function04')
            test_name = jsonpath.jsonpath(thread_list[1], '$..test_name')[0]
            testdate = thread_list[0]
            usecasenumber = 1
            callingTerminal_list = self.set_callingTerminal_list(test_name)
            # address_list = self.set_address_list(thread_list)
            # range_num = self.set_rangenum(test_name)
            fnum, tnum, x = self.set_phonenumber(test_name)
            work_book, work_sheet, save_file = self.set_save_file(thread_list)
            sence_num = 0
            row = 1  # 行
            c = 1  # 列
            for run_now in range(1, range_num):
                row += 1
                data = self.zzj_data(usecasenumber, test_name, testdate, fnum, tnum, callingTerminal_list, address_list,
                                     range_num, run_now, sence_num, lag_list=lag_list)
                for i in data:
                    work_sheet.cell(row=row, column=c, value=i)
                    # 设置居中
                    work_sheet.cell(row=row, column=c).alignment = Alignment(horizontal='center', vertical='center')
                    if i == '有杂音':
                        work_sheet.cell(row=row, column=c+2, value='主叫终端有杂音')  # --》 2~4次
                    if c == 14:
                        if run_now % 2 == 0:
                            work_sheet.merge_cells(start_row=run_now, start_column=1, end_row=run_now + 1, end_column=1)
                            work_sheet.merge_cells(start_row=run_now, start_column=2, end_row=run_now + 1, end_column=2)
                            usecasenumber += 1
                            sence_num += 2
                        c = 1
                    else:
                        c += 1
            #     pbar.update(10 * run_now + 1)
            #     # time.sleep(0.0001)
            # pbar.finish()
            self.set_border(work_sheet, 'N')
            work_book.save(save_file)
            # logger.info('{}保存成功'.format(test_name))
        except Exception as e:
            logger.error('调用save_function04异常：{}'.format(e))

    # 早拆线 保存方式
    @test_call
    def save_function05(self, thread_list, address_list, range_num, lag_list):
        # widgets = ['Progress: save_function05', Percentage(), ' ', Bar(marker=RotatingMarker('>-=')),
        #            ' ', ETA(), ' ', FileTransferSpeed()]
        # pbar = ProgressBar(widgets=widgets, maxval=100000).start()
        try:
            # logger.info('判断为【被叫早拆线】调用save_function05')
            test_name = jsonpath.jsonpath(thread_list[1], '$..test_name')[0]
            testdate = thread_list[0]
            usecasenumber = 1
            callingTerminal_list = self.set_callingTerminal_list(test_name)
            # address_list = self.set_address_list(thread_list)
            # range_num = self.set_rangenum(test_name)
            fnum, tnum, x = self.set_phonenumber(test_name)
            work_book, work_sheet, save_file = self.set_save_file(thread_list)
            sence_num = 0
            row = 1  # 行
            c = 1  # 列
            for run_now in range(1, range_num):
                row += 1
                data = self.zcx_data(usecasenumber, test_name, testdate, fnum, tnum, callingTerminal_list, address_list,
                                     range_num, run_now, sence_num, lag_list=lag_list)
                for i in data:
                    work_sheet.cell(row=row, column=c, value=i)
                    work_sheet.cell(row=row, column=c).alignment = Alignment(horizontal='center', vertical='center')
                    if c == 12:
                        if run_now % 2 == 0:
                            work_sheet.merge_cells(start_row=run_now, start_column=1, end_row=run_now + 1, end_column=1)
                            work_sheet.merge_cells(start_row=run_now, start_column=2, end_row=run_now + 1, end_column=2)
                            usecasenumber += 1
                            sence_num += 2
                        c = 1
                    else:
                        c += 1
            #     pbar.update(10 * run_now + 1)
            #     # time.sleep(0.0001)
            # pbar.finish()
            self.set_border(work_sheet, 'L')
            work_book.save(save_file)
            # logger.info('{}保存成功'.format(test_name))
        except Exception as e:
            logger.error('调用save_function05异常：{}'.format(e))

    # 呼叫等待保存
    @test_call
    def save_function06(self, thread_list, address_list, range_num, lag_list):
        # widgets = ['Progress: save_function06', Percentage(), ' ', Bar(marker=RotatingMarker('>-=')),
        #            ' ', ETA(), ' ', FileTransferSpeed()]
        # pbar = ProgressBar(widgets=widgets, maxval=100000).start()
        try:
            # logger.info('判断为【呼叫等待】调用save_function06')
            test_name = jsonpath.jsonpath(thread_list[1], '$..test_name')[0]
            testdate = thread_list[0]
            usecasenumber = 1
            # callingTerminal_list = self.set_callingTerminal_list(test_name)
            # address_list = self.set_address_list(thread_list)
            # range_num = self.set_rangenum(test_name)
            pnuma, pnumb, pnumc = self.set_phonenumber(test_name)
            work_book, work_sheet, save_file = self.set_save_file(thread_list)
            sence_num = 0
            row = 1  # 行
            c = 1  # 列
            for run_now in range(1, range_num):
                row += 1
                data = self.callwait_data(usecasenumber, test_name, testdate, pnuma, pnumb, pnumc, address_list,
                                          range_num, run_now, sence_num, lag_list=lag_list)
                for i in data:
                    work_sheet.cell(row=row, column=c, value=i)
                    work_sheet.cell(row=row, column=c).alignment = Alignment(horizontal='center', vertical='center')
                    if c == 12:
                        if run_now % 2 == 0:
                            work_sheet.merge_cells(start_row=run_now, start_column=1, end_row=run_now + 1, end_column=1)
                            work_sheet.merge_cells(start_row=run_now, start_column=2, end_row=run_now + 1, end_column=2)
                            usecasenumber += 1
                            sence_num += 2
                        c = 1
                    else:
                        c += 1
            #     pbar.update(10 * run_now + 1)
            #     # time.sleep(0.0001)
            # pbar.finish()
            self.set_border(work_sheet, 'L')
            work_book.save(save_file)
            # logger.info('{}保存成功'.format(test_name))
        except Exception as e:
             logger.error('调用save_function04异常：{}'.format(e))

    # 保存合集
    @test_call
    def save_data(self, thread_list):  # [time,{}]
        # logger.info('save_data-->thread_list:{}'.format(thread_list))
        test_name = jsonpath.jsonpath(thread_list[1], '$..test_name')[0]
        range_num = self.set_rangenum(test_name)
        lag_list = self.lag_list(test_name, range_num)
        # range_num = 30
        address_list = self.test_address_list(thread_list)  # -->list [d1.1,d1.2,d1.3,...]
        if test_name in ['无线弱覆盖', '无线重载']:
            self.save_function01(thread_list, address_list, range_num, lag_list)
            pass
        elif test_name == '基本呼叫':
            self.save_function02(thread_list, address_list, range_num, lag_list)
            pass
        elif test_name == '网络切换':
            self.save_function03(thread_list, address_list, range_num, lag_list)
            pass
        elif test_name == '被叫早摘机':
            self.save_function04(thread_list, address_list, range_num, lag_list)
            pass
        elif test_name == '被叫早拆线':
            self.save_function05(thread_list, address_list, range_num, lag_list)
        elif test_name in ['数据统计工作', '数据分析工作', '优化实施建议工作']:
            # logger.info(f'{test_name},跳过。。。。')
            pass
        else:  # 呼叫等待
            self.save_function06(thread_list, address_list, range_num, lag_list)

    # 多线程
    @test_call
    def more_line(self):
        """
        多线程入口
        :return:
        """
        start_time = time.strftime('%H:%M:%S', time.localtime())
        logger.info('someline启动线程')
        value = self.read_yaml()
        test_key_list = self.set_testname_list(value)  # 获取测试用例名称的集合 ['test_case01',...]
        # 通过测试用例名称集合的长度来确定启动线程池循环次数
        with ThreadPoolExecutor(max_workers=5) as pool:
            for i in range(len(test_key_list)):
                logger.info('第{}次循环执行{}用例'.format(i + 1, test_key_list[i]))
                # 需要当前测试用例的testdate列表集合来确定当前测试场景需要的线程数--testdate列表集合为启动线程的参数
                thread_list = self.set_thread_list(i, value, test_key_list)  # 取值 组装[[时间，测试集对应的值],[]]
                logger.info('第{}次循环thread_list--》：{}'.format(i + 1, thread_list))
                # 线程入口
                results = pool.map(self.save_data, thread_list)
                logger.info('-------------------------')
                for r in results:
                    logger.info(r)
        end_time = time.strftime('%H:%M:%S', time.localtime())
        # run_time = end_time - start_time
        logger.info(f'本次程序运行开始时间为：{start_time},结束时间为：{end_time}')


if __name__ == '__main__':
    # logger.add("media_optinization_log.log", rotation="10MB", encoding="utf-8", enqueue=True, compression="zip",
    #            retention="10 days")
    # now_path = os.getcwd()
    # yaml_file = os.path.join(now_path, 'jb_testcase.yaml')
    # dyaml_file = r'/PY/MiguTemplate/jb_call/debug_case.yaml'
    # save_file = r'D:/桌面/testcase/工单/2021基本呼叫工单/2021年5月基本呼叫/四川省端到端基本呼叫分析优化-{}({}).xlsx'
    # logger.add('./Log/log_{}.log'.format(time.strftime('%Y%m%d-%H%M%S', time.localtime())), rotation='20 MB',
    #            encoding='utf-8')
    # save_path : D:\桌面\testcase\工单\2021-8工单\基本呼叫

    m = JbCall()
    m.more_line()
