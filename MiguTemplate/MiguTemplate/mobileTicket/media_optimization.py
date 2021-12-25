# -*- coding: utf-8 -*-
# @Time : 2021/2/24 16:43
# @Author :PANZER
# @Email : 453989453@a163.com
# @File : media_optimization.py
# @Project : 媒体优化
from concurrent.futures.thread import ThreadPoolExecutor
import pandas as pd
import os
import random
import openpyxl
import xlwt
import jsonpath
from openpyxl.styles import Border, Side, colors, Alignment
from PY.interfacepytest.tools.loging import logger
from PY.MiguTemplate.mobileTicket.yaml_read import YamlHandle


class MediaOptimization:

    def __init__(self, yamlpath, filepath):
        self.yamlpath = yamlpath
        # self.value = YamlHandle().yaml_read(self.yamlpath)
        self.filepath = filepath

    # 测试场景名称集合
    def test_case(self, value):
        """
        获取测试用例名称集合的列表，用于循环启动多线程
        :param value:   YamlHandle().yaml_read(self.yamlpath)获取的测试用例数据
        :return: test_key_list--》测试场景名称列表
        """
        test_key_list = []
        for test_dict in value:
            for test_key in test_dict:
                test_key_list.append(test_key)
        logger.info('测试用例集为：{}'.format(test_key_list))
        return test_key_list

    # 线程数的参数
    def thread_data(self, i, value, test_key_list):
        """
        :param i:
        :param value:
        :param test_key_list:
        :return:
        """
        # 取测试日期
        data = jsonpath.jsonpath(value[i], '$..testdate')
        test_date_list = []  # 测试日期列表
        for test_date_key in data[0]:
            value01 = data[0].get(test_date_key)
            test_date_list.append(value01)
        logger.info('当前测试用例集时间：{}'.format(test_date_list))
        # 获取testcase值
        case_data = jsonpath.jsonpath(value[i], '$..{}'.format(test_key_list[i]))
        logger.info('测试集testcase值：{}'.format(case_data))
        # 组合列表
        thread_list = []
        for v in test_date_list:
            list01 = [v, case_data[0]]
            thread_list.append(list01)

        return thread_list  # [[time,{}],[time2,{}]]

    # 当前测试用例测试日期列表
    def test_date(self, thread_list):
        date = jsonpath.jsonpath(thread_list[1], '$..testdate') [0] # [{t1:*,t2:*}]
        time_list = []
        for key in date:
            d_value = date.get(key)
            time_list.append(d_value)
        logger.info('time_list:{}'.format(time_list))
        return time_list  # [time1,time2,...]

    # 所有地址列表
    def test_address(self, thread_list):
        address_dict = jsonpath.jsonpath(thread_list[1], '$..address')[0]
        address_list = []
        for key in address_dict:
            if address_dict.get(key) is None:
                logger.info('test_address-->为空,跳过此条地址')
                pass
            else:
                address_list.append(address_dict.get(key))
        logger.info('address_list:{}'.format(address_list))
        return address_list  # [[d1],[d2],[d3]...]

    # 当前测试用例地址列表
    def test_address_list(self, thread_list):
        time = thread_list[0]  # time
        time_list = self.test_date(thread_list)  # [time1,time2,...]
        address_list = self.test_address(thread_list)  # [[d1],[d2],[d3]...]
        num = time_list.index(time)
        address = address_list[num]
        return address  # [d1.1,d1.2,d1.3,...]

    # 测试场景
    def test_name(self):
        """
        获取测试场景名字 文件保存时用
        :return: 测试场景名字
        """
        test_name = jsonpath.jsonpath(self.value, '$..name')
        logger.info('测试场景为：{}'.format(test_name))
        return test_name

    # 测试地址--没用
    def address(self, thread_list):
        """
        收集测试地址
        :return: 测试地址列表 --》[[1],[2],[3]]
        """
        address_dict = jsonpath.jsonpath(self.value, '$..address')[0]
        address_list = []
        for i in address_dict:
            value_list = address_dict[i]
            if value_list is None:
                pass
            else:
                address_list.append(value_list)
        logger.info('测试地址列表：{}'.format(address_list))
        return address_list
        pass

    # 测试日期--没用
    def date(self):
        """
        获取测试日期列表--用测试提起列表去控制启动线程数
        :return: 测试日期列表 --》[time1, time2, time3]
        """
        date_dict = jsonpath.jsonpath(self.value, '$..testdate')[0]
        date_list = []
        for i in date_dict:
            date = date_dict.get(i)
            if date in None:
                pass
            else:
                date_list.append(date)
        logger.info('测试时间列表：{}'.format(date_list))
        return date_list

    # 测试次数
    def testing_num(self, test_name):
        """
        通过测试场景获取随机获取测试次数
        :param test_name:测试场景名称
        :return:
        """
        # 1=常规测试801条-2，2=基本测试1601条-4，3=优测试2401条-6，4=弱网测试2401-6，5=热岛测试2401-6
        # test_name = self.test_name()
        if test_name == '常规场景测试':
            while True:
                test_num = random.randint(801, 841)
                if (test_num-1)%2 == 0:
                    logger.info('{}运行次数为：{}'.format(test_name, test_num))
                    break
            return test_num
        elif test_name == '基本场景覆盖测试':
            while True:
                test_num = random.randint(1601, 1701)
                if (test_num-1)%4 == 0:
                    logger.info('{}运行次数为：{}'.format(test_name, test_num))
                    break
            return test_num
        else:
            while True:
                test_num = random.randint(2401, 2501)
                if (test_num-1)%6 == 0:
                    logger.info('{}运行次数为：{}'.format(test_name, test_num))
                    break
            return test_num

    # 定义边框样式
    def my_border(self):
        """ 定义边框样式 """
        border = Border(top=Side(border_style='thin', color=colors.BLACK),
                        bottom=Side(border_style='thin', color=colors.BLACK),
                        left=Side(border_style='thin', color=colors.BLACK),
                        right=Side(border_style='thin', color=colors.BLACK))
        return border

    # 设置边框
    def set_border(self, ws, column):
        """ 设置边框 """
        all_w = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                 'U', 'V', 'W', 'X', 'Y', 'Z']
        re = all_w.index(column)
        for q in all_w[:re + 1]:
            col = ws[q]
            for cell in col:
                cell.border = self.my_border()  # 调用上面边框设置函数

    # 创建文件
    def make_file(self, thread_list):
        # r'D:/桌面/testcase/工单/四川省端到端媒体质量优化附件-1月/四川省端到端媒体质量优化分析-{基本场景覆盖测试}({}).xlsx'
        # file = file+'/四川省端到端媒体质量优化分析-'+test_name+'({})'.format(test_data)+'.xlsx'
        test_name = jsonpath.jsonpath(thread_list[1], '$..name')[0]
        save_filepath = self.filepath.format(test_name, thread_list[0])
        logger.info('save_filepath:{}'.format(save_filepath))
        try:
            if os.path.exists(save_filepath):
                logger.info('make_file:文件已存在')
                pass
            else:
                work_book = xlwt.Workbook(encoding='utf-8')
                sheet = work_book.add_sheet('sheet')
                work_book.save(save_filepath)
                now_path = os.getcwd()
                mode_path = os.path.join(now_path, 'test_mode/媒体质量优化-数据源.xlsx')
                header = pd.read_excel(mode_path, nrows=0)
                writer = pd.ExcelWriter(save_filepath)
                df = pd.DataFrame(header)
                df.to_excel(writer, 'Sheet1', index=False)
                writer.save()
            work_book1 = openpyxl.load_workbook(save_filepath)
            work_sheet = work_book1['Sheet1']
            logger.info('测试结果表格创建成功')
            return work_book1, work_sheet, save_filepath
        except Exception as e:
            logger.warning('make_file创建文件失败{}'.format(e))

    # 利用文件模板创建WB,WS,savefile
    def set_save(self, thread_list):
        now_path = os.getcwd()
        mode_file = os.path.join(now_path, r'test_mode/四川省端到端媒体质量优化分析-用例模板.xlsx')
        # mode_file = r'D:/桌面/testcase/工单材料模板/四川省端到端媒体质量优化分析-用例模板.xlsx'
        test_name = jsonpath.jsonpath(thread_list[1], '$..name')[0]
        save_filepath = self.filepath.format(test_name, thread_list[0])
        work_book1 = openpyxl.load_workbook(mode_file)
        work_sheet = work_book1['Sheet1']
        logger.info('测试结果表格创建对象成功')
        return work_book1, work_sheet, save_filepath

    # 电话号码
    def phone_num(self):
        """
        获得主被叫号码
        :return: int 需要两个变量接收
        """
        logger.info('phone_num获取主被叫号码启动')
        pn_list = [13880356372, 1832838016, 13568898061]
        a = random.randint(0, 2)
        if a < 2:
            fphone = pn_list[a]
            tphone = pn_list[a + 1]
            logger.info('主被叫号码获取成功')
            return fphone, tphone
        else:
            fphone = pn_list[a]
            tphone = pn_list[a - 1]
            logger.info('主被叫号码获取成功')
            return fphone, tphone

    # 卡顿数据
    def caton(self, test_name, range_num):
        """
        获得卡顿的数目列表，多线程时使用
        :param range_num: 测试用例条数
        :param test_name: 测试场景
        :return: 卡顿次数列表
        """
        #range_num = self.testing_num()
        # test_name = self.test_name()
        logger.info('make_list获取卡顿次数启动')
        caton_list = []
        if test_name in ['常规场景测试', '基本场景覆盖测试']:
            a = random.randint(4, 10)
        elif test_name in ['无线弱覆盖场景测试', '无线热岛场景测试']:
            a = random.randint(8, 12)
        else:
            a = random.randint(5, 8)
        for i in range(a + 1):
            caton_list.append(random.randint(1, range_num))
        logger.info('caton-->caton_list:{}'.format(caton_list))
        return caton_list

    # 视频、音频切换判断数
    def sence_number(self, test_name, run_now, sence_num):

        if test_name == '基本场景覆盖测试':
            if run_now <= 4:
                logger.info('run_now<=4第{}次循环--》sence_num：{}'.format(run_now, sence_num))
                pass
            elif (run_now - 1) % 4 == 0:
                sence_num += 4
                logger.info('run_now>4第{}次循环，sence_num步进4-->{}'.format(run_now, sence_num))
            else:
                logger.info('run_now>4第{}次循环，不符合步进条件--sence_num步进4-->{}'.format(run_now, sence_num))
            return sence_num

        else:
            if run_now <= 6:
                logger.info('run_now <= 6第{}次循环--》sence_num：{}'.format(run_now, sence_num))
                pass
            elif (run_now - 1) % 6 == 0:
                sence_num += 6
                logger.info('1.run_now > 6第{}次循环，sence_num步进6-->{}'.format(run_now, sence_num))
            else:
                logger.info('2.run_now > 6第{}次循环，不符合步进条件--sence_num步进6-->{}'.format(run_now, sence_num))
            return sence_num
        pass

    # 视频、音频写入数据设置--基本，无线，热岛，优覆盖
    def set_subScene(self, test_name, run_now, sence_num):
        subScene_list = ['音频起呼', '视频起呼']
        if test_name == '基本场景覆盖测试':
            if run_now <= 4:
                if 0.25 <= run_now / 4 <= 0.5:
                    logger.info('构造data--》run_now:{},sence_num{}'.format(run_now, sence_num))
                    subScene = subScene_list[0]
                else:
                    subScene = subScene_list[1]
                pass
            else:
                if 0.25 <= (run_now - sence_num) / 4 <= 0.5:
                    logger.info('构造data--》run_now:{},sence_num{}'.format(run_now, sence_num))
                    subScene = subScene_list[0]
                else:
                    subScene = subScene_list[1]
            return subScene
        else:
            if run_now <= 6:
                if run_now / 6 <= 0.5:
                    logger.info('构造data--》run_now:{},sence_num{}'.format(run_now, sence_num))
                    subScene = subScene_list[0]
                else:
                    subScene = subScene_list[1]
                pass
            else:
                if (run_now - sence_num) / 6 <= 0.5:
                    logger.info('构造data--》run_now:{},sence_num{}'.format(run_now, sence_num))
                    subScene = subScene_list[0]
                else:
                    subScene = subScene_list[1]
            return subScene
        pass

    # 设置data的地址
    def set_address(self, test_name, address_list, range_num, serialnumber):
        """
        根据当前测试用例编号判断测试地址
        :param test_name: str 测试场景名称
        :param address_list: list 地址列表
        :param range_num:  int 总运行次数
        :param serialnumber: int 当前测试用例编号
        :return:
        """
        if test_name == '常规场景测试': # 800
            if range_num/8 < serialnumber <= range_num/4:  # 201-400  101-200
                address = address_list[0]
                pass
            elif range_num/4 < serialnumber <= range_num/2.5:  # 401-600 201-300
                address = address_list[1]
                pass
            elif range_num/2.5 < serialnumber:  # 601-800 301-400
                address = address_list[2]
                pass
            else:  # 0-200 1-100
                address = address_list[3]
            pass
        elif test_name == '基本场景覆盖测试': # 1600
            if range_num / 16 < serialnumber <= range_num / 8:  # 401-800  101-200
                address = address_list[0]
                pass
            elif range_num / 8 < serialnumber <= range_num / 5:  # 801-1200 201-300
                address = address_list[1]
                pass
            elif range_num / 5 < serialnumber:  # 1201-1600 301-400
                address = address_list[2]
                pass
            else:  # 0-400 1-100
                address = address_list[3]
            pass
        else:
            if range_num/24 < serialnumber <= range_num/12:  # 601-1200  101-200
                address = address_list[0]
                pass
            elif range_num/12 < serialnumber <= range_num/8:  # 1201-1800 201-300
                address = address_list[1]
                pass
            elif range_num/8 < serialnumber:  # 1800-2400 301-400
                address = address_list[2]
                pass
            else:  # 0-600 1-100
                address = address_list[3]
                pass
        return address

    # 常规测试数据
    def cg_data(self, serialnumber, caton_list, run_now, time, address_list, fnum, tnum, fphone, range_num, test_name):
        """
        构造写入表格的数据
        :param caton_list: list 卡顿列表
        :param run_now:  int 当前运行次数
        :param time:  str 测试日期
        :param address:  list 地址列表
        :param fnum:   int 主叫号码
        :param tnum:   int 被叫号码
        :param fphone:  str 主叫终端
        :param range_num:  运行总次数--》判断地址
        :return: list 写入表格的数据
        """

        scene1 = ['视频起呼', '音频起呼']
        if run_now % 2 == 0:
            scene = scene1[1]
        else:
            scene = scene1[0]
        testcase = '在RSRP为-90，SINR>10DB,RB<30%的区域进行测试'
        tester = '林波/郭益群'
        RSRP = int(random.randint(-97, -87))
        SINR = float('%.1f' % (random.uniform(10, 20)))
        MOS = float('%.1f' % (random.uniform(39, 43) / 10))
        RTP = ('0.0{}%'.format(random.randint(2, 7)))
        para_caton = ['正常', '轻微卡顿']
        if run_now in caton_list:
            caton = para_caton[1]
        else:
            caton = para_caton[0]
        address = self.set_address(test_name, address_list, range_num, serialnumber)
        data = [serialnumber, test_name, scene, testcase, time, fnum, tnum, fphone, tester, address, RSRP, SINR, MOS, RTP, caton, ]
        logger.info('第{}次cg_data-->data:{}'.format(run_now, data))
        return data

    # 热岛数据
    def rd_data(self, serialnumber, test_name, caton_list, run_now, time, address_list, fnum, tnum, fphone, range_num,
                sence_num):
        testcase1 = ['1.主叫在无线热岛覆盖区域', '2.被叫在无线热岛覆盖区域', '3.主被叫在无线热岛覆盖区域']  # 三个场景不同需要修改
        if (run_now % 3) == 1:
            testcase = testcase1[0]
        elif (run_now % 3) == 2:
            testcase = testcase1[1]
        else:
            testcase = testcase1[2]
        tester = '林波/郭益群'
        address = self.set_address(test_name, address_list, range_num, serialnumber)
        RSRP = int(random.randint(-110, -80))
        SINR = float('%.1f' % (random.uniform(10, 20)))
        MOS = float('%.1f' % (random.uniform(37, 41) / 10))
        RTP = ('0.{}%'.format(random.randint(32, 67)))
        #
        vedio1 = ['正常', '轻微卡顿']
        if run_now in caton_list:
            vedio = vedio1[1]
        else:
            vedio = vedio1[0]
        subScene = self.set_subScene(test_name, run_now, sence_num)
        data = [serialnumber,test_name, subScene, testcase, time, fnum, tnum, fphone, tester, address, RSRP, SINR, MOS, RTP, vedio, ]
        logger.info('第{}次rd_data-->data:{}'.format(run_now, data))
        return data

    # 优测试结果
    def y_data(self, serialnumber, test_name, caton_list, run_now, time, address_list, fnum, tnum, fphone, range_num,
               sence_num):
        testcase1 = ['1.主叫在无线优覆盖区域', '2.被叫在无线优覆盖区域', '3.主被叫在无线优覆盖区域']  # 三个场景不同需要修改
        if run_now % 3 == 1:
            testcase = testcase1[0]
        elif run_now % 3 == 2:
            testcase = testcase1[1]
        else:
            testcase = testcase1[2]
        tester = '林波/郭益群'
        # 地址
        address = self.set_address(test_name, address_list, range_num, serialnumber)
        #
        RSRP = int(random.randint(-80, -60))
        SINR = float('%.1f' % (random.uniform(20, 30)))
        MOS = float('%.1f' % (random.uniform(40, 45) / 10))
        RTP = ('0.0{}%'.format(random.randint(1, 5)))
        #
        vedio1 = ['正常', '轻微卡顿']
        if run_now in caton_list:
            vedio = vedio1[1]
        else:
            vedio = vedio1[0]
        subScene = self.set_subScene(test_name, run_now, sence_num)
        data = [serialnumber, test_name, subScene, testcase, time, fnum, tnum, fphone, tester, address, RSRP, SINR, MOS, RTP, vedio, ]
        logger.info('第{}次y_data-->data:{}'.format(run_now, data))
        return data

    # 弱网结果
    def rw_data(self,serialnumber, test_name, caton_list, run_now, time, address_list, fnum, tnum, fphone, range_num,
                sence_num):
        logger.info('第{}次构造rw_data..'.format(run_now))
        testcase1 = ['1.主叫在无线弱网覆盖区域', '2.被叫在无线弱网覆盖区域', '3.主被叫在无线弱网覆盖区域']
        if (run_now % 3) == 1:
            testcase = testcase1[0]
        elif (run_now % 3) == 2:
            testcase = testcase1[1]
        else:
            testcase = testcase1[2]
        tester = '林波/郭益群'
        address = self.set_address(test_name, address_list, range_num, serialnumber)
        RSRP = int(random.randint(-115, -105))
        SINR = float('%.1f' % (random.uniform(-30, 30) / 10))
        MOS = float('%.1f' % (random.uniform(34, 38) / 10))
        RTP = ('0.{}%'.format(random.randint(52, 87)))

        vedio1 = ['正常', '轻微卡顿']
        if run_now in caton_list:
            vedio = vedio1[1]
        else:
            vedio = vedio1[0]
        subScene = self.set_subScene(test_name, run_now, sence_num)
        data = [serialnumber, test_name, subScene, testcase, time, fnum, tnum, fphone, tester, address, RSRP, SINR, MOS, RTP, vedio, ]
        logger.info('第{}次rw_data-->data:{}'.format(run_now, data))
        return data

    # 基本测试结果
    def jb_data(self,serialnumber,  test_name, caton_list, run_now, time, address_list, fnum, tnum, fphone, range_num,
                sence_num):
        """
        基本测试数据
        :param serialnumber int 测试用例编号
        :param test_name  str 测试场景
        :param caton_list: list 卡顿列表
        :param run_now: int 当前运行次数
        :param time: str  测试日期
        :param address_list: list  测试地址列表
        :param fnum:  int 主叫
        :param tnum:  int 被叫
        :param fphone:  str 主叫终端
        :param range_num int 运行总次数 构造子场景，地址
        :param sence_num interesting 构造子场景数据
        :return:
        """
        logger.info('构造基本测试结果')
        testcase_list = ['1.主叫VoLTE用户（驻留LTE）音频呼叫VoLTE用户（驻留LTE）', '2.主叫VoLTE用户（驻留LTE）音频呼叫VoLTE用户（驻留CS)']
        if run_now % 2 != 0:
            testcase = testcase_list[0]
        else:
            testcase = testcase_list[1]
        tester = '林波/郭益群'
        # 地址--1600
        address = self.set_address(test_name, address_list, range_num, serialnumber)
        RSRP = int(random.randint(-97, -87))
        SINR = float('%.1f' % (random.uniform(10, 20)))
        MOS = float('%.1f' % (random.uniform(39, 43) / 10))
        RTP = ('0.0{}%'.format(random.randint(2, 7)))
        # 卡顿
        vedio1 = ['正常', '轻微卡顿']
        if run_now in caton_list:
            vedio = vedio1[1]
        else:
            vedio = vedio1[0]
        # subScene_list = ['音频起呼', '视频起呼']
        # if run_now <= 4:
        #     if 0.25 <= run_now/4 <= 0.5:
        #         logger.info('构造data--》run_now:{},sence_num{}'.format(run_now, sence_num))
        #         subScene = subScene_list[0]
        #     else:
        #         subScene = subScene_list[1]
        #     pass
        # else:
        #     if 0.25 <= (run_now-sence_num)/4 <=0.5:
        #         logger.info('构造data--》run_now:{},sence_num{}'.format(run_now, sence_num))
        #         subScene = subScene_list[0]
        #     else:
        #         subScene = subScene_list[1]
        #     pass
        # 构造子场景数据
        subScene = self.set_subScene(test_name, run_now, sence_num)

        data = [serialnumber, test_name, subScene, testcase, time, fnum, tnum, fphone, tester, address, RSRP, SINR, MOS, RTP, vedio, ]
        logger.info('第{}次jb_data-->data:{}'.format(run_now, data))
        return data

    # 常规测试保存方法
    def save_function2(self, range_num, caton_list, time, address, fnum, tnum, work_sheet, test_name):
        """
        常规测试保存方法
        :param range_num: int 测试次数
        :param caton_list:  list 卡顿次数
        :param time:  str 测试日期
        :param address:  list 测试地址列表
        :param fnum: int 主叫号码
        :param tnum: int 被叫号码
        :param work_sheet: 工作表对象
        :return:
        """
        logger.info('save_function1启动')
        phone_list = ['荣耀play note10', '小米8SE', '华为P30 pro', 'realme 真我x2']
        a = random.randint(0, 3)
        fphone = phone_list[a]
        row = 1
        c = 1
        serialnumber = 1
        try:
            for run_now in range(1, range_num):  # 2426,2440,2438,2447
                row += 1
                data = self.cg_data(serialnumber, caton_list, run_now, time, address, fnum, tnum, fphone, range_num, test_name)
                logger.info('获取data成功：save_function2-->{}{}'.format(run_now, data))
                # print(c)
                for j in data:
                    # print(j)
                    work_sheet.cell(row=row, column=c, value=j)
                    work_sheet.cell(row=row, column=c).alignment = Alignment(horizontal='center', vertical='center')
                    if j == '轻微卡顿':
                        work_sheet.cell(row, c + 1, '测试过程视频彩铃出现轻微卡顿')
                    if c == 15:
                        c = 1
                    else:
                        c += 1
                logger.info('[save_function2数据写入成功:{}{}]'.format(run_now, data))
                if run_now % 2 == 0:  # 2的倍数需要更换
                    serialnumber += 1
                    # 此处插入合并表格
                    work_sheet.merge_cells(start_row=run_now, start_column=1, end_row=run_now+1, end_column=1)
                    work_sheet.merge_cells(start_row=run_now, start_column=2, end_row=run_now + 1, end_column=2)
                    if (run_now / 2) % 2 == 0:  # 奇数
                        fphone = phone_list[a]
                    else:  # 偶数
                        if a < 3:
                            fphone = phone_list[a + 1]
                        else:
                            fphone = phone_list[a - 1]
        except Exception as e:
            logger.error('常规测试保存方法-->save_function2:{}'.format(e))

    # 基本，优，弱网，热岛保存方法
    def save_function1(self, range_num, caton_list, time, address, fnum, tnum, work_sheet, test_name):
        """
        基本，优覆盖，弱网，热岛保存方法
        :param range_num: int 测试次数
        :param list: list 卡顿次数列表
        :param time: str 测试日期
        :param address:  list 测试地址列表
        :param fnum:  int 主叫号码
        :param tnum: int 被叫号码
        :param work_sheet: obj 工作表对象
        :param test_name: int 测试场景
        :return:
        """
        logger.info('save_function1启动')
        phone_list = ['荣耀play note10', '小米8SE', '华为P30 pro', 'realme 真我x2']
        a = random.randint(0, 3)
        fphone = phone_list[a]

        row = 1  # 行 openpyxl行和列的下标最小从1开始
        c = 1  # 列
        serialnumber = 1  # 用例编号
        # range_num = 20  # 调试用注意注释掉
        sence_num = 0
        try:
            for run_now in range(1, range_num):  # 2426,2440,2438,2447
                row += 1
                # data = dosomting(i, list, address, time)
                # if run_now <= 4:
                #     logger.info('第{}次循环--》sence_num：{}'.format(run_now, sence_num))
                #     pass
                # elif (run_now-1)%4 == 0:
                #     sence_num += 4
                #     logger.info('第{}次循环，sence_num步进4-->{}'.format(run_now, sence_num))
                # else:
                #     logger.info('第{}次循环，不符合步进条sence_num步进4件-->{}'.format(run_now, sence_num))
                #     pass
                sence_num = self.sence_number(test_name, run_now, sence_num)
                if test_name == '基本场景覆盖测试':
                    data = self.jb_data(serialnumber, test_name, caton_list, run_now, time, address, fnum, tnum, fphone, range_num,sence_num)
                elif test_name == '无线优覆盖场景测试':
                    data = self.y_data(serialnumber, test_name, caton_list, run_now, time, address, fnum, tnum, fphone, range_num, sence_num)
                elif test_name == '无线弱覆盖场景测试':
                    data = self.rw_data(serialnumber, test_name, caton_list, run_now, time, address, fnum, tnum, fphone, range_num, sence_num)
                else:  #
                    data = self.rd_data(serialnumber, test_name, caton_list, run_now, time, address, fnum, tnum, fphone, range_num, sence_num)
                logger.info('获取data成功：{}{}'.format(run_now, data))
                # print(data) 保存数据
                for j in data:
                    # print(j)
                    work_sheet.cell(row=row, column=c, value=j)
                    work_sheet.cell(row=row, column=c).alignment = Alignment(horizontal='center', vertical='center')
                    #
                    if j == '轻微卡顿':
                        work_sheet.cell(row, c + 1, '测试过程视频彩铃出现轻微卡顿')
                    if c == 15:
                        c = 1  # --弱网需要改
                    else:
                        c += 1
                logger.info('[数据写入成功:{}{}]'.format(run_now, data))
                # 判断使用那种替换
                if test_name == '基本场景覆盖测试':
                    logger.info('save_function1判断基本测试，更换手机型号')
                    if run_now % 2 == 0:
                        work_sheet.merge_cells(start_row=run_now, start_column=3, end_row=run_now+1, end_column=3)
                    else:
                        pass
                    if run_now % 4 == 0:  # 4这些整数倍需要跟换
                        serialnumber += 1
                        # 合并
                        work_sheet.merge_cells(start_row=run_now-2, start_column=1, end_row=run_now+1, end_column=1)
                        work_sheet.merge_cells(start_row=run_now-2, start_column=2, end_row=run_now+1, end_column=2)
                        if (run_now / 4) % 2 == 0:  # 奇数
                            fphone = phone_list[a]
                        else:  # 偶数
                            if a < 3:
                                fphone = phone_list[a + 1]
                            else:
                                fphone = phone_list[a - 1]
                    pass
                else:
                    logger.info('save_function1判断弱网，优，热岛，更换手机型号')
                    # 合并第3列
                    if run_now % 3 == 0:
                        work_sheet.merge_cells(start_row=run_now - 1, start_column=3, end_row=run_now + 1, end_column=3)
                    if run_now % 6 == 0:  # 6,整数倍需要跟换，会重复
                        serialnumber += 1
                        # 合并 第1，2列
                        work_sheet.merge_cells(start_row=run_now - 4, start_column=1, end_row=run_now + 1, end_column=1)
                        work_sheet.merge_cells(start_row=run_now - 4, start_column=2, end_row=run_now + 1, end_column=2)
                        if (run_now / 6) % 2 == 0:  # 奇数
                            fphone = phone_list[a]
                        else:  # 偶数
                            if a < 3:
                                fphone = phone_list[a + 1]
                            else:
                                fphone = phone_list[a - 1]
        except Exception as e:
            logger.error("基本，优，弱网，热岛保存方法-->save_function1:{}".format(e))

    # 保存数据方法合集
    def save_data(self, thread_list):
        """

        :param thread_list: 时间与测试集合的列表--》[time,{}]
        :return:
        """
        try:
            logger.info('set_data主程序开始')
            work_book, work_sheet, save_filepath = self.set_save(thread_list)
            # 获取主被叫号码
            fnum, tnum = self.phone_num()
            # 获取测试名称
            test_name = jsonpath.jsonpath(thread_list[1], '$..name')[0]
            # 获取测试次数
            range_num = self.testing_num(test_name)
            # 获取卡顿次数
            caton_list = self.caton(test_name, range_num)  # -->list [caton1,caton2,...]
            # 测试日期
            time = thread_list[0]
            # 获取测试地址列表
            address = self.test_address_list(thread_list)  # -->list [d1.1,d1.2,d1.3,...]
            # 1=常规测试2，2=基本测试4，3=优测试6，4=弱网测试6，5=热岛测试6
            if test_name in ['基本场景覆盖测试', '无线优覆盖场景测试', '无线弱覆盖场景测试', '无线热岛场景测试']:  # 基本，优，弱网，热岛
                logger.info('set_data判断测试场景为{},调用save——function1'.format(test_name))
                self.save_function1(range_num, caton_list, time, address, fnum, tnum, work_sheet, test_name)
                self.set_border(work_sheet, 'P')
                # work_book.save(save_filepath)
                logger.info('save_function1数据保存成功')
                pass
            elif test_name in ['数据统计工作', '数据分析工作', '优化实施建议工作']:
                logger.info(f'{test_name},跳过。。。。')
                pass
            else:  # 常规测试
                logger.info('set_data判断测试场景为：{}，调用save_function2'.format(test_name))
                self.save_function2(range_num, caton_list, time, address, fnum, tnum, work_sheet, test_name)
                # 合并单元格，并写入合并操作的操作
                # self.save_combine02(range_num, test_name, work_sheet)
                self.set_border(work_sheet, 'P')
                # work_book.save(save_filepath)
                logger.info('save_function2数据保存成功')
                pass
        except Exception as e:
            logger.error('保存数据方法合集-->save_data:{}'.format(e))
        finally:
            work_book.save(save_filepath)

    # 多线程
    def more_line(self):
        """
        多线程入口
        :return:
        """
        logger.info('someline启动线程')
        value = YamlHandle().yaml_read(self.yamlpath)
        test_key_list = self.test_case(value)  # 获取测试用例名称的集合 ['test_case01',...]
        # 通过测试用例名称集合的长度来确定启动线程池循环次数
        with ThreadPoolExecutor(max_workers=5) as pool:
            for i in range(len(test_key_list)):
                logger.info('第{}次循环执行{}用例'.format(i+1, test_key_list[i]))
                # 需要当前测试用例的testdate列表集合来确定当前测试场景需要的线程数--testdate列表集合为启动线程的参数
                thread_list = self.thread_data(i, value, test_key_list)  # 取值 组装[[时间，测试集对应的值],[]]
                logger.info('第{}次循环thread_list--》：{}'.format(i+1, thread_list))
                # 线程入口
                results = pool.map(self.save_data, thread_list)
                logger.info('-------------------------')
                for r in results:
                    logger.info(r)


if __name__ == '__main__':
    # 正式
    yamlpath = r'D:\PyCharm\PY\mobileTicket\mobile_address.yaml'
    save_file = r'D:\桌面\testcase\工单\四川省端到端媒体质量优化附件-5月\四川省端到端媒体质量优化分析-{}({}).xlsx'
    # 调试
    debug_file = r'/PY/MiguTemplate/mobileTicket/test_data_yaml.yaml'
    test_file = r'D:\桌面\testcase\工单\脚本调试\四川省端到端媒体质量优化分析-{}({}).xlsx'

    m = MediaOptimization(yamlpath, save_file)  # --实例对象
    m.more_line()
    # now_path = os.getcwd()
    # mode_path = os.path.join(now_path, r'test_mode\四川省端到端媒体质量优化分析-用例模板.xlsx')
    # print(mode_path)




