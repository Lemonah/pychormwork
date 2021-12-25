# -*- coding: utf-8 -*-
# @Time : 2021/7/13 16:29
# @Author :PANZER
# @Email : 453989453@a163.com
# @File : COMPATIBILITY_TEST.py 兼容测试
# @Project : PyCharm

import os, openpyxl, jsonpath, yaml, random, json
import time, sys
import tqdm
from loguru import logger
from functools import partial
from openpyxl.styles import Border, Side, colors, Alignment
from concurrent.futures.thread import ThreadPoolExecutor
from tqdm import tqdm


def test_call(func):
    """
    接口调用记录
    :param func: 装饰的函数
    :return:
    """

    def inner(*args, **kwargs):
        logger.info(f"开始调用函数：{func.__name__}")
        res = func(*args, **kwargs)
        logger.info(f"结束调用函数：{func.__name__}")
        return res

    return inner


class CompatibilityTest:

    yam_file = os.path.join(os.getcwd(), 'compatibility_test.yaml')
    # save_file = input('输入保存文件的路径：')


    def my_border(self):
        """ 定义边框样式 """
        border = Border(top=Side(border_style='thin', color=colors.BLACK),
                        bottom=Side(border_style='thin', color=colors.BLACK),
                        left=Side(border_style='thin', color=colors.BLACK),
                        right=Side(border_style='thin', color=colors.BLACK))
        return border

        # 设置边框


    def set_border(self, ws, column):
        """ 设置边框 """
        all_w = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                 'U', 'V', 'W', 'X', 'Y', 'Z']
        re = all_w.index(column)
        for q in all_w[:re + 1]:
            col = ws[q]
            for cell in col:
                cell.border = self.my_border()  # 调用上面边框设置函数


    def handle_yaml(self):
        with open(self.yam_file, encoding='utf-8') as f:
            yaml_value = yaml.load(f, Loader=yaml.FullLoader)
            # logger.info(f'yaml文件值为:{yaml_value}')
            return yaml_value


    def set_save_path_f(self, yaml_value, save_path):
        date = jsonpath.jsonpath(yaml_value, '$..date')
        test_name = jsonpath.jsonpath(yaml_value, '$..test_name')[0]
        # logger.info(f'test_name = {test_name}')
        if test_name == '分辨率测试':
            mode_file = os.path.join(os.getcwd(), '业务综合优化-终端兼容性测试（分辨率测试）mode.xlsx')
            save_file_path = os.path.join(save_path, f'业务综合优化-终端兼容性测试（分辨率测试）({date[0]}).xlsx')
        elif test_name == '呼叫转移':
            mode_file = os.path.join(os.getcwd(), '业务综合优化-终端兼容性测试（呼叫转移）mode.xlsx')
            save_file_path = os.path.join(save_path, f'业务综合优化-终端兼容性测试（呼叫转移）({date[0]}).xlsx')
        else:
            mode_file = os.path.join(os.getcwd(), '业务综合优化-终端兼容性测试（网络切换）mode.xlsx')
            save_file_path = os.path.join(save_path, f'业务综合优化-终端兼容性测试（网络切换）({date[0]}).xlsx')
        work_book = openpyxl.load_workbook(mode_file)
        work_sheet = work_book['Sheet1']
        # logger.info(f'文件保存地址为：{save_file_path}')
        return work_book, work_sheet, save_file_path

    def set_resolvingPower(self):
        """
        分辨率设置,写两次
        :return:
        """
        resolvingPower_list = ['1024*576', '1024*768', '1280*960', '540*960', '576*1024', '640*680',
                               '960*540', '800*800']
        # resolvingPower = random.sample(resolvingPower_list, 6)
        return resolvingPower_list

    def set_averageBitrate(self):
        """
        设置平均码率，写2次
        :return:
        """
        averageBitrate_list = [670, 673, 675, 678, 679, 680, 682, 684, 685, 686, 691, 697, 694,
                               700, 701, 702, 710, 713, 714, 720, 721, 723, 724, 727, 730, 735, 738, 739, 740, 741,
                               745, 755, 758, 759, 715, 760, 762, 763, 767, 768, 774, 782, 783, 787, 788, 790, 792, 799,
                               794, 795]
        # averageBitrate = random.sample(averageBitrate_list, 10)
        return averageBitrate_list

    def set_peakBitRate(self):
        """
        设置峰值码率，写两次
        :return:
        """
        peakBitRate_list = [803, 804, 808, 813, 826, 827, 828, 829, 834, 836, 837, 839, 840, 845, 847, 849, 851, 852,
                            854, 856, 857, 860, 861, 866, 868, 870, 878, 880, 884, 891, 893, 894, 902, 905, 908, 909,
                            911, 912, 917, 918, 935, 938, 941, 945, 947, 950, 954, 958, 851, 862, 865, 873, 897, 922,
                            929, 938, 957]
        # peakBitRate = random.sample(peakBitRate_list, 10)
        return peakBitRate_list

    def set_call_phone(self):
        call_phone_list_all = ['oppo R11','vivo iQOO','荣耀9x','华为P30 pro','小米6x','三星Galaxy s8','realme 真我x2',
                            'VIVO s1 pro','OPPO Reno ACE','HUAWEI MATE10 pro','小米8SE','VIVO Y75',
                            '荣耀play note10','红米6','HUAWEI MATE9','小米mix2','华为nova 4','小米cc9','红米 k30 ','VIVO y66i',
                           'oppo A7x','VIVO y85','华为P20 pro','荣耀10','oppo R17 pro','oppo FIND X','vivo x23',
                           '荣耀v20','小米8青春版','小米9','小米 Play','vivo x27','vivo x2liA','vivo NEX','HUAWEI MATE 30']
        call_phone_list = random.sample(call_phone_list_all, 5)
        return call_phone_list

    def set_call_num(self):
        """
        设置主叫号码3个写两遍
        :return:
        """
        call_num_all = [15982094112, 18382722728, 13637932712, 13568898061, 18328380168, 15172533008]
        call_num_list = random.sample(call_num_all, 3)
        return call_num_list

    def set_called_num(self):
        """"
        被叫号码列表
        """
        called_num_all = [15202876373, 15708457282, 18202877366, 15928798398, 18202877366, 18302804461]
        called_num_list = random.sample(called_num_all, 3)
        return called_num_list

    def set_rsrp(self):
        rsrp = int(random.randint(-110, -80))
        return rsrp

    def set_sinr(self):
        sinr = float('%.1f' % random.uniform(1, 26))
        return sinr

    def set_lag_list(self):
        """
        轻微卡顿列表 25次
        :return:
        """
        lag_list = []
        lag_num = random.randint(25, 35)
        for i in range(1, lag_num + 1):
            lag_list.append(random.randint(1, 601))
        # logger.info(f'轻微卡顿次数为：{lag_list}')
        return lag_list

    @test_call
    def set_than_lag_list(self, lag_list):
        """
        中度卡顿列表,是轻微卡顿的三分之一，循环之前要确定
        :return:
        """
        than_lag_list = []
        num = len(lag_list)
        than_lag_num = int(num*0.3)
        for i in range(1, than_lag_num):
            lag_num = random.randint(1, 601)
            if lag_num in lag_list:
                # logger.info('当前数字在轻微卡顿列表中跳过')
                continue
            else:
                than_lag_list.append(lag_num)
        # logger.info(f'中度卡顿次数为：{than_lag_list}')
        return than_lag_list

    def set_play_ring(self, run_now, lag_list, than_lag_list):
        play_ring_list = ['正常', '轻微卡顿', '中度卡顿']
        if run_now in lag_list:
            return play_ring_list[1]
        elif run_now in than_lag_list:
            return play_ring_list[2]
        else:
            return play_ring_list[0]

    def set_problem(self, run_now, lag_list, than_lag_list):
        problem_list = ['无', '视频彩铃播放中出现很轻微的卡顿', '视频彩铃播放中出现中度卡顿']
        if run_now in lag_list:
            return problem_list[1]
        elif run_now in than_lag_list:
            return problem_list[2]
        else:
            return problem_list[0]

    def set_ring__relation(self, ring_list, resolvingPower_list, averageBitrate_list, peakBitRate_list):
        """
        将铃音与分辨率，平均码率，峰值码率一一对应
        :param ring_list:
        :param resolvingPower_list:
        :param averageBitrate_list:
        :param peakBitRate_list:
        :return:
        """
        resolvingPower_dict = {}
        averageBitrate_dict = {}
        peakBitRate_dict = {}
        for ring in ring_list:
            resolvingPower_dict[ring] = random.choice(resolvingPower_list)
            averageBitrate_dict[ring] = random.choice(averageBitrate_list)
            peakBitRate_dict[ring] = random.choice(peakBitRate_list)
        return resolvingPower_dict, averageBitrate_dict, peakBitRate_dict

    def exl_data01(self):
        """
        分辨率测试数据
        """
        ring_name = 1  # 3 铃声名字  是一个列表
        resolvingPower = ['1024*576', '1024*768', '1280*960', '540*960', '576*1024', '640*680', '960*540']  # 5
        averageBitrate = [679,685,694,701,715,762,763,790,794,795]  # 7平均码率
        peakBitRate = [851,862,865,873,897,922,929,938,957]  # 8峰值码率
        test_date = 1  # 9 测试日期
        call_phone = []  # 10主叫机型
        call_num = []  # 11主叫号码
        rsrp = 1  # 13随机
        sinr = 1  # 14随机
        play_ring = ['正常', '轻微卡顿', '中度卡顿']  # 16
        problem = ['无', '视频彩铃播放中出现很轻微的卡顿', '视频彩铃播放中出现中度卡顿']  # 17
    # _______________________分辨率______________________________
    @test_call
    def save_function01(self, yaml_value, work_sheet):
        """
        9,13,14,16,17一起写
        :param yaml_value:
        :return:
        """
        try:
            date = jsonpath.jsonpath(yaml_value, '$..date')[0]
            lag_list = self.set_lag_list()
            than_lag_list = self.set_than_lag_list(lag_list)
            row = 2
            for run_now in tqdm(range(1, 601)):
                rsrp = self.set_rsrp()
                sinr = self.set_sinr()
                play_ring = self.set_play_ring(run_now, lag_list, than_lag_list)
                problem = self.set_problem(run_now, lag_list, than_lag_list)
                work_sheet.cell(row=row, column=9, value=date)
                work_sheet.cell(row=row, column=13, value=rsrp)
                work_sheet.cell(row=row, column=14, value=sinr)
                work_sheet.cell(row=row, column=16, value=play_ring)
                work_sheet.cell(row=row, column=17, value=problem)
                row += 1
                time.sleep(0.1)
        except Exception as e:
            logger.error(f"调用save_function01异常:{e}")

    @test_call
    def save_function02(self, work_sheet):
        """
        ,10,11
        :param yaml_value:
        :param work_sheet:
        :return:
        """
        try:
            call_phont_list = self.set_call_phone()
            call_num_list = self.set_call_num()
            row = 2
            for run_now in tqdm(range(1, 301)):  # 301
                # resolvingPower = self.set_resolvingPower()
                # averageBitrate = self.set_averageBitrate()
                # peakBitRate = self.set_peakBitRate()
                call_phone = random.choice(call_phont_list)
                call_num = random.choice(call_num_list)
                for i in range(2):
                    # work_sheet.cell(row=row, column=5, value=resolvingPower)
                    # work_sheet.cell(row=row, column=7, value=averageBitrate)
                    # work_sheet.cell(row=row, column=8, value=peakBitRate)
                    work_sheet.cell(row=row, column=10, value=call_phone)
                    work_sheet.cell(row=row, column=11, value=call_num)
                    row += 1
                # row += 2
                time.sleep(0.1)
        except Exception as e:
            logger.error(f"调用save_function02异常:{e}")

    @test_call
    def save_function03(self, yaml_value, work_sheet):
        """
        3列,5,7,8
        :return:
        """
        try:
            ring_list = jsonpath.jsonpath(yaml_value, '$..r1')[0]  # -->[r1,r2,r3...]
            resolvingPower_list = self.set_resolvingPower()  # ---> [6]
            averageBitrate_list = self.set_averageBitrate()  # ---> [10]
            peakBitRate_list = self.set_peakBitRate()        # ---> [10]
            resolvingPower_dict, averageBitrate_dict, peakBitRate_dict = self.set_ring__relation(ring_list,
                                                                                                 resolvingPower_list,
                                                                                                 averageBitrate_list,
                                                                                                 peakBitRate_list)
            row = 2
            for run in tqdm(range(30)):
                for ring in ring_list:
                    for i in range(2):
                        work_sheet.cell(row=row, column=3, value=ring)  # 铃音名字
                        work_sheet.cell(row=row, column=5, value=resolvingPower_dict[ring])  # 分辨率
                        work_sheet.cell(row=row, column=7, value=averageBitrate_dict[ring])  # 平均码率
                        work_sheet.cell(row=row, column=8, value=peakBitRate_dict[ring])  # 峰值码率
                        row += 1
                time.sleep(0.1)
        except Exception as e:
            logger.error(f"调用save_function03异常: {e}")
    # _______________________网络切换和呼叫转移_______________________________

    @test_call
    def save_function04(self, yaml_value, work_sheet):
        """
        写4，10，11，13，14
        :param yaml_value:
        :param work_sheet:
        :return:
        """
        try:
            date = jsonpath.jsonpath(yaml_value, '$..date')[0]
            test_name = jsonpath.jsonpath(yaml_value, '$..test_name')[0]
            lag_list = self.set_lag_list()
            than_list = self.set_than_lag_list(lag_list)
            if test_name == '网络切换':
                range_num = 600
            else:
                range_num = 900
            row = 2
            for run_now in tqdm(range(range_num)):
                rsrp = self.set_rsrp()
                sinr = self.set_sinr()
                play_ring = self.set_play_ring(run_now, lag_list, than_list)
                problem = self.set_problem(run_now, lag_list, than_list)
                work_sheet.cell(row=row, column=4, value=date)
                work_sheet.cell(row=row, column=10, value=rsrp)
                work_sheet.cell(row=row, column=11, value=sinr)
                work_sheet.cell(row=row, column=13, value=play_ring)
                work_sheet.cell(row=row, column=14, value=problem)
                row += 1
                time.sleep(0.1)
        except Exception as e:
            logger.error(f"调用save_function04异常: {e}")

    @test_call
    def save_function05(self, yaml_value, work_sheet):
        """
        写5，6，7列
        :param work_sheet:
        :return:
        """
        try:
            test_name = jsonpath.jsonpath(yaml_value, '$..test_name')[0]
            call_phone_list = self.set_call_phone()
            call_num_list = self.set_call_num()
            called_num_list = self.set_called_num()
            if test_name == '网络切换':
                range_num01 = 300
                range_num02 = 2
            else:
                range_num01 = 300
                range_num02 = 3
            row = 2
            for run_now in tqdm(range(range_num01)):
                call_phone = random.choice(call_phone_list)
                call_num = random.choice(call_num_list)
                called_num = random.choice(called_num_list)
                for i in range(range_num02):
                    work_sheet.cell(row=row, column=5, value=call_phone)
                    work_sheet.cell(row=row, column=6, value=call_num)
                    work_sheet.cell(row=row, column=7, value=called_num)
                    row += 1
                time.sleep(0.1)
        except Exception as e:
            logger.error(f"调用save_function05异常: {e}")

    @test_call
    def save_function06(self, yaml_value, work_sheet):
        try:
            address_list = jsonpath.jsonpath(yaml_value, '$..l1')[0]
            test_name = jsonpath.jsonpath(yaml_value, '$..test_name')[0]
            if test_name == '网络切换':
                range_num01 = 75
                range_num02 = 2
            else:
                range_num01 = 100
                range_num02 = 3
            row = 2
            for run_now in tqdm(range(range_num01)):
                for address in address_list:
                    for i in range(range_num02):
                        work_sheet.cell(row=row, column=9, value=address)
                        row += 1
                time.sleep(0.1)
        except Exception as e:
            logger.error(f"调用save_function06异常: {e}")
    # _______________________________________________________

    def save_resolvingPower(self, yaml_value, save_path):
        """
        分辨率测试保存方法
        :return:
        """
        try:
            # yaml_value = self.handle_yaml()
            work_book, work_sheet, save_file = self.set_save_path_f(yaml_value, save_path)
            self.save_function01(yaml_value, work_sheet)
            self.save_function02(work_sheet)
            self.save_function03(yaml_value, work_sheet)
            work_book.save(save_file)
            # logger.info(f'{jsonpath.jsonpath(yaml_value,"$..test_name")}工单保存成功')
        except Exception as e:
            logger.error(f'保存方法异常：{e}')


    def save_networkSwitching(self, yaml_value, save_path):
        """
        网络切换保存
        :return:
        """
        try:
            # yaml_value = self.handle_yaml()
            work_book, work_sheet, save_file = self.set_save_path_f(yaml_value, save_path)
            self.save_function04(yaml_value,  work_sheet)
            self.save_function05(yaml_value, work_sheet)
            self.save_function06(yaml_value, work_sheet)
            work_book.save(save_file)
            pass
        except Exception as e:
            logger.error(f'网络切换保存异常：{e}')
    # __________________________________________________________

    def save_all(self, yaml_value, save_path):
        test_name = jsonpath.jsonpath(yaml_value, '$..test_name')[0]
        if test_name == '分辨率测试':
            self.save_resolvingPower(yaml_value, save_path)
        else:
            self.save_networkSwitching(yaml_value, save_path)

    def set_data_all(self, yaml_value):

        data_all = []
        for test_data in yaml_value:
            test_name = jsonpath.jsonpath(test_data, '$..test_name')[0]  # -->列表套列表需要去除外层列表
            if test_name == '分辨率测试':
                num = 1
                ring_name_list = jsonpath.jsonpath(test_data, '$..ring_name')[0]
                for ring_name in ring_name_list.values():
                    data_dict01 = {'testcase01': {'test_name': '分辨率测试', 'ring_name': {'r1': ''}, 'time': {'date': ''}}}
                    data_dict01['testcase01']['ring_name']['r1'] = ring_name
                    data_dict01['testcase01']['time']['date'] = jsonpath.jsonpath(test_data, f'$..date{num}')[0]
                    print('分辨率测试：', data_dict01)
                    data_all.append(data_dict01)
                    num += 1
            else:
                num = 1
                address_list = jsonpath.jsonpath(test_data, '$..address')[0]
                for address in address_list.values():
                    data_dict02 = {
                        'testcase02': {'test_name': f'{test_name}', 'address': {'l1': []}, 'time': {'date': ''}}}
                    data_dict02['testcase02']['address']['l1'] = address
                    data_dict02['testcase02']['time']['date'] = jsonpath.jsonpath(test_data, f'$..date{num}')[0]
                    data_all.append(data_dict02)
                    num += 1
        return data_all

    # --------------------------------------------------------

    @test_call
    def more_thread(self):
        save_path = input(r'请输入《业务综合测试》文件保存路径：')
        time.sleep(3)
        logger.info('---------多线程启动----------')
        yaml_value = self.handle_yaml()
        data_all = self.set_data_all(yaml_value)
        with ThreadPoolExecutor(max_workers=10) as pool:
            results = pool.map(partial(self.save_all, save_path=save_path), data_all)
            logger.info('-------------------------')
            for r in results:
                logger.info(r)
        logger.info('------------多线程结束------------')


if __name__ == '__main__':
    # D:\桌面\testcase\工单\脚本调试
    c = CompatibilityTest()
    # v = c.handle_yaml()  # 列表
    # c.set_data_all(v)
    #print(v)
    c.more_thread()


