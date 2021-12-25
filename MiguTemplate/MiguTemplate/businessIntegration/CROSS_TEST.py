# -*- coding: utf-8 -*-
# @Time : 2021/7/13 16:29
# @Author :PANZER
# @Email : 453989453@a163.com
# @File : CROSS_TEST.py 交叉测试
# @Project : PyCharm

import os, openpyxl, jsonpath, yaml, random, time
from tqdm import tqdm
from loguru import logger
from functools import partial
from openpyxl.styles import Border, Side, colors, Alignment
from concurrent.futures.thread import ThreadPoolExecutor
from PY.MiguTemplate.businessIntegration.COMPATIBILITY_TEST import test_call


class CrossTest(object):

    mode_file = os.path.join(os.getcwd(), '业务综合优化-厂商交叉测试 - mode.xlsx')


    # 表格写入位置 3~13

    # 目前思路不用这个方法
    def set_call_fun(self, run_now):
        """
        设置起呼方式
        :param run_now: 当前运行次数
        :return:
        """
        call_fun_list = ['视频起呼', '音频起呼']
        if run_now % 2 == 0:
            call_fun = call_fun_list[1]
        else:
            call_fun = call_fun_list[0]
        return call_fun

    def my_border(self):
        """ 定义边框样式 """
        border = Border(top=Side(border_style='thin', color=colors.BLACK),
                        bottom=Side(border_style='thin', color=colors.BLACK),
                        left=Side(border_style='thin', color=colors.BLACK),
                        right=Side(border_style='thin', color=colors.BLACK))
        return border

        # 设置边框

    def set_border(self, ws, column):
        """ 设置边框 """
        all_w = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                 'U', 'V', 'W', 'X', 'Y', 'Z']
        re = all_w.index(column)
        for q in all_w[:re + 1]:
            col = ws[q]
            for cell in col:
                cell.border = self.my_border()  # 调用上面边框设置函数

    # 设置保存文件
    def save_file(self, filepath, yaml_value):
        try:
            # logger.info('设置保存文件')
            city_name = jsonpath.jsonpath(yaml_value, '$..city_name')[0]
            date = jsonpath.jsonpath(yaml_value, '$..date')
            work_book = openpyxl.load_workbook(self.mode_file)
            work_sheet = work_book['Sheet1']
            save_file = os.path.join(filepath, f'业务综合优化-厂商交叉测试 - {city_name}({date[0]}).xlsx')
            # logger.info(f'文件保存路径为：{save_file}')
            return work_book, work_sheet, save_file
        except Exception as e:
            logger.error(f'设置保存文件异常：{e}')
            raise e

    @test_call
    def hanlde_yaml(self):
        try:
            yaml_file_path = os.path.join(os.getcwd(), 'CROSS_DATA.yaml')
            with open(yaml_file_path, encoding='utf-8') as f:
                yaml_value = yaml.load(f, Loader=yaml.FullLoader)
                # logger.info(f'读取到的yaml为：{yaml_value}')
                return yaml_value
        except Exception as e:
            logger.error(f'yaml文件读取失败：{e}')

    def set_call_phone(self):
        """
        设置主叫终端，5个,循环前确定，每循环两次更换一次
        :return:
        """
        call_phone_list01 = ['华为mate30', '荣耀9x', 'VIVO X20', 'Redmi Note8 Pro', 'oppo R11', '小米10', '华为P30 pro',
                           'VIVO Y75', 'oppo R17 pro']
        call_phone_list = random.sample(call_phone_list01, 5)
        # logger.info(f'主叫终端列表为：{call_phone_list}')
        return call_phone_list

    def set_called_phone(self):
        """
        循环前确定
        :return:
        """
        called_phone_list01 = ['小米8', '小米6x', '荣耀v20', '华为P20', 'VIVO x20', 'oppo FIND X']
        called_phone = random.choice(called_phone_list01)
        # logger.info(f'被叫终端为：{called_phone}')
        return called_phone

    def set_phone_num(self):
        """
        主叫号码两个，循环前确定，每两次换一个号码，循环之前确定
        :return:
        """
        phone_num_list01 = ['18302804461', '15202876373', '15708457282', '18202877366', '15928798398']
        phone_num_list = random.sample(phone_num_list01,2)
        # logger.info(f'主叫号码为：{phone_num_list}')

    # 轻微卡顿列表
    def set_lag_list(self):
        """
        轻微卡顿列表,25~35 次，循环之前确定
        :return: list
        """
        lag_list = []
        lag_num = random.randint(25, 35)
        for i in range(1, lag_num+1):
            lag_list.append(random.randint(1, 800))
        # logger.info(f'轻微卡顿次数为：{lag_list}')
        return lag_list

    # 中度卡顿列表
    def set_than_lag_list(self, lag_list):
        """
        中度卡顿列表,是轻微卡顿的三分之一，循环之前要确定
        :return:
        """
        than_lag_list = []
        num = len(lag_list)
        than_lag_num = int(num*0.3)
        for i in range(1, than_lag_num):
            lag_num = random.randint(1, 800)
            if lag_num in lag_list:
                # logger.info('当前数字在轻微卡顿列表中跳过')
                continue
            else:
                than_lag_list.append(lag_num)
        # logger.info(f'中度卡顿次数为：{than_lag_list}')
        return than_lag_list

    # 设置最后一列的值
    def set_Problem(self, run_now, lag_list, than_lag_list):
        """
        将当前运行次数与卡顿列表的对比确定最后一列‘问题描述的’的值,循环内确定
        :param run_now: 当前运行次数
        :param lag_list: 轻微卡顿列表
        :param than_lag_list: 中度卡顿列表
        :return:
        """
        problem_list = ['无', '视频彩铃播放中出现中度卡顿', '视频彩铃播放中出现很轻微的卡顿']
        if run_now in lag_list:
            return problem_list[2]
        elif run_now in than_lag_list:
            return problem_list[1]
        else:
            return problem_list[0]

    # 设置播放效果是否正常
    def set_play_ring(self, run_now, lag_list, than_lag_list):
        """
        确定铃音播放情况列的值，循环内确定
        :param run_now:
        :param lag_list:
        :param than_lag_list:
        :return:
        """
        play_ring = ['正常', '中度卡顿', '轻微卡顿']
        if run_now in lag_list:
            return play_ring[2]
        elif run_now in than_lag_list:
            return play_ring[1]
        else:
            return play_ring[0]

    # 设置 rsrp，sinr，是否播放预期彩铃值
    def set_three_list(self):
        """
        设置 rsrp，sinr，是否播放预期彩铃值,循环内确定 每次值不一样
        :return:
        """
        rsrp = int(random.uniform(-110, -80))
        sinr = float('%.1f' % (random.uniform(3, 26)))
        extced_ring = '是'
        three_lsit = [rsrp, sinr, extced_ring]
        return three_lsit

    # 设置主叫号码
    def set_call_num_list(self):
        call_num_list_all = [18302804461, 15202876373, 15708457282, 18202877366, 15928798398]
        call_num_list = random.sample(call_num_list_all, 2)
        return call_num_list

    def exl_data(self):

        # serialNumber = '定义一个，逢二递增'
        # call_fun = ['视频起呼', '音频起呼']
        # testScenarios = '同一核心网下不同厂商间的网络视频彩铃兼容测试'
        testdate = '固定日期，yaml中读取'  # 4
        call_phone = ['华为mate30', '荣耀9x', 'VIVO X20', 'Redmi Note8 Pro', 'oppo R11', '小米10', '华为P30 pro']  # 5
        call_num = ['18302804461', '15202876373', '15708457282', '18202877366', '15928798398']  # 6
        called_phone = ['小米8', '小米6x', '荣耀v20', '华为P20', 'VIVO x20', 'oppo FIND X']         # 7
        tester = '林昀陟/唐安发'  # 8
        test_address = 'yaml'  # 9
        rsrp = random.uniform(-110, -80)  # 10
        sinr = float('%.1f' % (random.uniform(3, 26)))  # 11
        extced_ring = '是'  # 12
        play_ring = 'caton_list*2'  # 13
        Problem = ['无', '视频彩铃播放中出现中度卡顿', '视频彩铃播放中出现很轻微的卡顿']  # 14

    # ----------------------------写入方法---------------------------------
    # 4日期，7被叫终端，8测试员，13铃声播放情况，14 问题描述列一起写 循环800次
    def save_function01(self, yaml_value, work_sheet,):
        """"
        保存文件的方法
        """
        # widgets = ['Progress: save_function01', Percentage(), ' ', Bar(marker=RotatingMarker('>-=')),
        #            ' ', ETA(), ' ', FileTransferSpeed()]
        # pbar = ProgressBar(widgets=widgets, maxval=10000).start()
        try:
            test_date = jsonpath.jsonpath(yaml_value, '$..date')[0]  # 测试日期
            called = self.set_called_phone()  # 被叫终端
            tester = '林昀陟/唐安发'
            lag_list = self.set_lag_list()
            than_lag_list = self.set_than_lag_list(lag_list)
            row = 1  # 行
            for run_now in tqdm(range(1, 801)):
                row += 1  # 从第二行开始写
                work_sheet.cell(row=row, column=4, value=test_date)
                work_sheet.cell(row=row, column=7, value=called)
                work_sheet.cell(row=row, column=8, value=tester)
                work_sheet.cell(row=row, column=13, value=self.set_play_ring(run_now, lag_list, than_lag_list))
                work_sheet.cell(row=row, column=14, value=self.set_Problem(run_now, lag_list, than_lag_list))
            #     pbar.update(10 * run_now + 1)
                time.sleep(0.0001)
            # pbar.finish()
        except Exception as e:
            logger.error(f"调用save_function01异常{e}")

    # 10，11，12 列一起写 循环800次
    def save_function02(self, work_sheet):
        # widgets = ['Progress: save_function02', Percentage(), ' ', Bar(marker=RotatingMarker('>-=')),
        #            ' ', ETA(), ' ', FileTransferSpeed()]
        # pbar = ProgressBar(widgets=widgets, maxval=10000).start()
        try:
            row = 1
            col = 10
            for run_now in tqdm(range(1, 801)):
                three_list = self.set_three_list()
                row += 1
                for data in three_list:
                    work_sheet.cell(row=row, column=col, value=data)
                    if col == 12:
                        col = 10
                    else:
                        col += 1
                # pbar.update(10 * run_now + 1)
            time.sleep(0.0001)
            # pbar.finish()
        except Exception as e:
            logger.error(f"调用save_function02异常{e}")

    # 5，6，列一起写 循环 400次
    def save_function03(self, work_sheet):
        # widgets = ['Progress: save_function03', Percentage(), ' ', Bar(marker=RotatingMarker('>-=')),
        #            ' ', ETA(), ' ', FileTransferSpeed()]
        # pbar = ProgressBar(widgets=widgets, maxval=10000).start()
        try:
            call_phone_list = self.set_call_phone()  # 5主叫终端列表
            call_num_list = self.set_call_num_list()  # 6主叫号码
            row = 2
            for run_now in tqdm(range(1, 401)):  # 401
                call_phone = random.choice(call_phone_list)
                call_num = random.choice(call_num_list)
                for i in range(2):
                    work_sheet.cell(row=row, column=5, value=call_phone)  # 5
                    work_sheet.cell(row=row, column=6, value=call_num)  # 6
                    row += 1
                # pbar.update(10 * run_now + 1)
                time.sleep(0.0001)
            # pbar.finish()
        except Exception as e:
            logger.error(f"调用save_function03异常{e}")

    # 9 地址
    def save_function04(self, yaml_value, work_sheet):
        # widgets = ['Progress: save_function04', Percentage(), ' ', Bar(marker=RotatingMarker('>-=')),
        #            ' ', ETA(), ' ', FileTransferSpeed()]
        # pbar = ProgressBar(widgets=widgets, maxval=10000).start()
        try:
            address_list = jsonpath.jsonpath(yaml_value, '$..l')[0]  # 9地址列表
            row = 2
            for run_now in tqdm(range(1, 101)):  # 101
                for address in address_list:
                    for i in range(2):
                        work_sheet.cell(row=row, column=9, value=address)
                        row += 1
                # pbar.update(10 * run_now + 1)
                time.sleep(0.0001)
            # pbar.finish()
        except Exception as e:
            logger.error(f"调用save_function04异常{e}")
    # ------------------------------------------------------

    # 美化居中
    @test_call
    def beautify(self, work_sheet):
        # row = 2
        for row in range(2, 801):
            for col in range(4, 15):
                work_sheet.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')

    # 保存方法汇总--用于多线程
    @test_call
    def save_all(self, yaml_value, filepath):
        """
        这里的yaml_value是已经解包后的字典
        :param filepath:
        :param yaml_value:
        :return:
        """
        try:
            work_book, work_sheet, save_path = self.save_file(filepath, yaml_value)
            self.save_function01(yaml_value, work_sheet)
            self.save_function02(work_sheet)
            self.save_function03(work_sheet)
            self.save_function04(yaml_value, work_sheet)
            self.beautify(work_sheet)
            self.set_border(work_sheet, 'N')
            work_book.save(save_path)
        except Exception as e:
            logger.error(f'save_all error:{e}')
        finally:
            work_book.save(save_path)

    def set_data_all(self, yaml_value):
        data_all = []
        for test_data in yaml_value:
            city_name_list = jsonpath.jsonpath(test_data, '$..city_name')[0]  # -->列表套列表需要去除外层列表
            num = 1
            for city_name in city_name_list.values():
                data_dict01 = {'testcase01': {'city_name': f'{city_name}', 'address': {'l': ''}, 'time': {'date': ''}}}
                data_dict01['testcase01']['address']['l'] = jsonpath.jsonpath(test_data, f'$..l{num}')[0]
                data_dict01['testcase01']['time']['date'] = jsonpath.jsonpath(test_data, f'$..date{num}')[0]
                # print('厂商交叉测试：', data_dict01)
                data_all.append(data_dict01)
                num += 1
        return data_all

    # ----------------------------------------

    # 多线程
    @test_call
    def more_thread(self):
        filepath = input(r'请输入<业务综合-厂商交叉测试>的路径：')
        time.sleep(3)
        yaml_value = self.hanlde_yaml()
        data_all = self.set_data_all(yaml_value)
        with ThreadPoolExecutor(max_workers=10) as pool:
            results = pool.map(partial(self.save_all, filepath=filepath), data_all)
            logger.info('-------------------------')
            for r in results:
                logger.info(r)


if __name__ == '__main__':
    # D:\桌面\testcase\工单\脚本调试
    c = CrossTest()
    # v = c.hanlde_yaml()
    # add = jsonpath.jsonpath(v, '$..l')[0]
    # print(add)
    # three = c.set_three_list()
    # print(three)
    c.more_thread()